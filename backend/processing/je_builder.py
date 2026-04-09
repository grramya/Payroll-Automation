# =============================================================================
# processing/je_builder.py — Assemble and export the Journal Entry file
# =============================================================================
"""
Builds the final QBO-format Journal Entry DataFrame from the processed lines,
adds the balancing credit entry, and exports to Excel.

The output file layout exactly matches "JE for Payroll.xlsx":
  Row 1  : "Organisation Name" label in column D
  Row 5  : Company name in column A
  Row 6  : Column headers (multi-line, matching QBO import template exactly)
  Rows 7-8 : Blank spacer rows
  Row 9+ : Data rows
"""

from __future__ import annotations

import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import re
from config import JE_COLUMNS, JE_DESCRIPTION_ORDER

# GL account for the balancing credit entry
_ACCRUED_PAYROLL_ACCOUNT = "Accrued Expenses:Accrued Payroll"

# Exact column headers as they appear in the reference JE file (multi-line = \n)
_EXCEL_HEADERS = [
    "Post?",
    "Journal Number",
    "Entry Date",
    "Journal Description\none Narration per MJ is enough",
    "Account\nSelect from list",
    "Account ID",
    "Customer\nSelect from list",
    "Vendor\nSelect from list",
    "Employee\nSelect from list",
    "Location\nSelect from list",
    "Class\nSelect from list",
    "Tax Rate\nSelect from list",
    "Tax Application ON",
    "Currency\nSelect from list",
    "Debit\n(exc. Tax)",
    "Credit\n(exc. Tax)",
    "Adjustment",
    "QBO Edit ID\n(do not modify or delete)",
]

# Maps our internal DataFrame column names → Excel column position (1-based)
_COL_POSITION = {
    "Post?":             1,
    "Journal Number":    2,
    "Entry Date":        3,
    "Journal Description": 4,
    "Account":           5,
    "Account ID":        6,
    "Customer":          7,
    "Vendor":            8,
    "Employee":          9,
    "Location":          10,
    "Class":             11,
    "Tax Rate":          12,
    "Tax Application ON": 13,
    "Currency":          14,
    "Debit (exc. Tax)":  15,
    "Credit (exc. Tax)": 16,
    "Adjustment":        17,
    "QBO Edit ID":       18,
}


# ---------------------------------------------------------------------------
# Sorting helper
# ---------------------------------------------------------------------------

def _extract_pay_date(journal_number: str, entry_date: str) -> str:
    """Extract MM/DD/YYYY pay date from journal_number (e.g. 'Salary for 03/15/2026').
    Falls back to entry_date if no date pattern found."""
    m = re.search(r'\d{1,2}/\d{1,2}/\d{4}', journal_number)
    return m.group(0) if m else entry_date


def _sort_key(line: dict) -> tuple:
    """
    Returns a sort tuple so JE lines appear in the same order as Book3:
      Salary → Tech Stipend → Bonus/Commission → Separation Pay →
      Health Insurance → Admin Fees → Payroll Taxes → 401k →
      Reimbursements → Provision
    """
    desc = line.get("Journal Description", "")
    dept = line.get("Class", "") or ""

    # Strip date suffix for matching: "Salary for 03/15/2026" → "Salary"
    base = re.sub(r'\s+for\s+\d{1,2}/\d{1,2}/\d{4}$', '', desc).strip()
    # Also strip "COGS - " prefix for COGS variants
    base_no_cogs = re.sub(r'^(COGS - |COS - )', '', base).strip()

    if "Separation Pay" in desc:
        return (JE_DESCRIPTION_ORDER.get("Separation Pay", 30), desc, dept)
    if desc.startswith("Sales Commission for "):
        return (JE_DESCRIPTION_ORDER.get("Commission", 25), desc, dept)
    if desc.startswith("Reimb "):
        return (JE_DESCRIPTION_ORDER.get("Reimbursement", 100), desc, dept)

    # Try exact match, then base (no date), then base without COGS prefix
    for key in (desc, base, base_no_cogs):
        if key in JE_DESCRIPTION_ORDER:
            return (JE_DESCRIPTION_ORDER[key], desc, dept)

    return (500, desc, dept)


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_je(
    regular_lines: list[dict],
    special_lines: list[dict],
    journal_number: str,
    entry_date: str,
    provision_description: str = "",
    company_name: str = "Concertiv, Inc.",
) -> pd.DataFrame:
    """
    Assemble all JE lines into a QBO-ready DataFrame.

    Parameters
    ----------
    regular_lines : list[dict]
        Department-level aggregated lines from aggregator.aggregate_by_department()
    special_lines : list[dict]
        Employee-level lines from aggregator.process_special_columns()
    journal_number : str
        e.g. "Salary for 01/15/2025"
    entry_date : str
        e.g. "01/15/2025"
    provision_description : str
        Label for the balancing credit row.
        Defaults to "Provision for {entry_date}".
    company_name : str
        Written to the file header (row 5).

    Returns
    -------
    pd.DataFrame  (also carries company_name as an attribute for export use)
    """
    # Extract pay date for date-aware descriptions (e.g. "Salary for 03/15/2026")
    pay_date = _extract_pay_date(journal_number, entry_date)

    # Descriptions that get a pay-date suffix appended
    _DATE_SUFFIX_DESCS = {"Salary", "401k", "COGS - 401k"}

    all_lines = regular_lines + special_lines
    all_lines.sort(key=_sort_key)

    rows = []
    total_debit = 0.0
    total_credit = 0.0

    for line in all_lines:
        debit  = float(line.get("Debit",  0) or 0)
        credit = float(line.get("Credit", 0) or 0)
        total_debit  += debit
        total_credit += credit

        raw_desc = line.get("Journal Description", "")
        # Append pay date to Salary and 401k descriptions (Book3 format)
        if raw_desc in _DATE_SUFFIX_DESCS:
            final_desc = f"{raw_desc} for {pay_date}"
        else:
            final_desc = raw_desc

        # Location = Class (identical in Book3 for all rows)
        class_val = line.get("Class") or None

        rows.append({
            "Post?":               "Yes",
            "Journal Number":      journal_number,
            "Entry Date":          entry_date,
            "Journal Description": final_desc,
            "Account":             line.get("Account", ""),
            "Account ID":          line.get("Account ID") or None,
            "Customer":            None,
            "Vendor":              line.get("Vendor")  or None,
            "Employee":            None,           # Book3: Employee col always blank
            "Location":            None,
            "Class":               class_val,
            "Tax Rate":            None,
            "Tax Application ON":  None,
            "Currency":            "USD",
            "Debit (exc. Tax)":    debit  if debit  > 0 else None,
            "Credit (exc. Tax)":   credit if credit > 0 else None,
            "Adjustment":          "No",
            "QBO Edit ID":         None,
        })

    # Balancing credit entry (Provision)
    net = round(total_debit - total_credit, 2)
    if abs(net) >= 0.01:
        prov_label = provision_description.strip() if provision_description else ""
        if not prov_label:
            prov_label = f"Provision for {entry_date}"

        rows.append({
            "Post?":               "Yes",
            "Journal Number":      journal_number,
            "Entry Date":          entry_date,
            "Journal Description": prov_label,
            "Account":             _ACCRUED_PAYROLL_ACCOUNT,
            "Account ID":          None,
            "Customer":            None,
            "Vendor":              None,
            "Employee":            None,
            "Location":            None,           # Provision has no dept/class
            "Class":               None,
            "Tax Rate":            None,
            "Tax Application ON":  None,
            "Currency":            "USD",
            "Debit (exc. Tax)":    None,
            "Credit (exc. Tax)":   net,
            "Adjustment":          "No",
            "QBO Edit ID":         None,
        })

    je_df = pd.DataFrame(rows, columns=JE_COLUMNS)
    # Carry company_name for use in export
    je_df.attrs["company_name"] = company_name
    return je_df


# ---------------------------------------------------------------------------
# Export — replicates the exact layout of "JE for Payroll.xlsx"
# ---------------------------------------------------------------------------

def _build_workbook(je_df: pd.DataFrame) -> Workbook:
    """
    Create an openpyxl Workbook that exactly matches the reference JE structure:
      Row 1-4 — hidden metadata rows (Row 1 has "Organisation Name" in D1)
      Row 5   — Company name in A5 (green, bold)
      Row 6   — Column headers (dark green fill, white bold text, wrap)
      Row 7   — hidden spacer (height 2.25)
      Row 8   — empty spacer
      Row 9+  — Data rows
    """
    company_name = je_df.attrs.get("company_name", "Concertiv, Inc.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Styles ──────────────────────────────────────────────────────────────
    FONT_NAME      = "Aptos Narrow"
    GREEN_FILL     = PatternFill("solid", fgColor="19973D")     # header bg
    WHITE_BOLD_10  = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    GREEN_BOLD_12  = Font(name=FONT_NAME, size=12, bold=True, color="19973D")
    GREEN_BOLD_10  = Font(name=FONT_NAME, size=10, bold=True, color="19973D")  # "Yes" cells
    NORMAL_10      = Font(name=FONT_NAME, size=10)
    WRAP_CENTER    = Alignment(wrap_text=True, vertical="center", horizontal="center")
    WRAP_LEFT      = Alignment(wrap_text=True, vertical="center", horizontal="left")
    CENTER         = Alignment(vertical="center", horizontal="center")
    LEFT           = Alignment(vertical="center", horizontal="left")
    RIGHT_ALIGN    = Alignment(vertical="center", horizontal="right")

    # ── Column widths: auto-fit based on content, with min/max guardrails ──────
    # Will be applied after all cells are written (see end of function)
    COL_MIN_WIDTH = 8.0
    COL_MAX_WIDTH = 50.0

    # ── Hidden rows (rows 1-4) ────────────────────────────────────────────────
    for r in [1, 2, 3, 4]:
        ws.row_dimensions[r].hidden = True

    # Row heights
    ws.row_dimensions[1].height = 43.5
    for r in [2, 3, 4]:
        ws.row_dimensions[r].height = 36.75
    ws.row_dimensions[5].height = 26.25
    # rows 6 and 7 set below in header section

    # ── Row 1: "Organisation Name" label (hidden) ────────────────────────────
    ws.cell(row=1, column=4, value="Organisation Name")

    # ── Row 5: Company name ───────────────────────────────────────────────────
    cell_co = ws.cell(row=5, column=1, value=company_name)
    cell_co.font      = GREEN_BOLD_12
    cell_co.alignment = LEFT

    # ── Row 6: Stars-only row (thin, same green, ★ top-left in required cols) ──
    STAR_COLS   = {1, 2, 3, 5, 15, 16}   # Post?, Journal#, Entry Date, Account, Debit, Credit
    STAR_FONT   = Font(name=FONT_NAME, size=8, bold=True, color="FFFFFF")
    TOP_LEFT    = Alignment(vertical="top", horizontal="left")

    ws.row_dimensions[6].height = 13
    for col_idx in range(1, 18):
        cell = ws.cell(row=6, column=col_idx)
        cell.fill = GREEN_FILL
        if col_idx in STAR_COLS:
            cell.value     = "★"
            cell.font      = STAR_FONT
            cell.alignment = TOP_LEFT

    # ── Row 7: Main header text row (centered, bold, white on green) ──────────
    ws.row_dimensions[7].height = 28
    for col_idx, header_text in enumerate(_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header_text)
        cell.fill      = GREEN_FILL
        cell.font      = WHITE_BOLD_10
        cell.alignment = WRAP_CENTER

    # ── Row 8: hidden spacer  |  Row 9: empty spacer ─────────────────────────
    ws.row_dimensions[8].hidden = True
    ws.row_dimensions[8].height = 2.25
    ws.row_dimensions[9].height = 11.25

    # ── Row 10+: data rows ────────────────────────────────────────────────────
    DATA_START_ROW = 10
    for row_offset, (_, data_row) in enumerate(je_df.iterrows()):
        excel_row = DATA_START_ROW + row_offset
        ws.row_dimensions[excel_row].height = 20.25

        for col_name, col_pos in _COL_POSITION.items():
            val = data_row.get(col_name)
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                cell = ws.cell(row=excel_row, column=col_pos, value=val)
                cell.font = NORMAL_10

                if col_name in ("Debit (exc. Tax)", "Credit (exc. Tax)"):
                    cell.alignment     = RIGHT_ALIGN
                    cell.number_format = '#,##0.00'
                elif col_name in ("Post?", "Currency", "Adjustment"):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT

    # ── Auto-fit column widths based on actual cell content ───────────────────
    for col_cells in ws.iter_cols():
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is None:
                continue
            # For multi-line headers, measure the longest line only
            lines = str(cell.value).split("\n")
            cell_len = max(len(line) for line in lines)
            if cell_len > max_len:
                max_len = cell_len
        # Scale character count → Excel width units (approx 1.2 chars per unit)
        fitted = max_len * 1.2 + 2
        ws.column_dimensions[col_letter].width = max(COL_MIN_WIDTH, min(fitted, COL_MAX_WIDTH))

    return wb


def export_je_to_bytes(je_df: pd.DataFrame) -> bytes:
    """
    Serialize the JE to an in-memory Excel file and return raw bytes.
    Used for Streamlit download buttons.
    """
    wb = _build_workbook(je_df)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def export_je_to_file(je_df: pd.DataFrame, output_path: str) -> str:
    """
    Write the JE to an Excel file on disk matching the reference layout.
    Returns the output path.
    """
    wb = _build_workbook(je_df)
    wb.save(output_path)
    return output_path
