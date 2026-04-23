"""
Base P&L (Class) (Individual) generator.

Exactly mirrors 'Base P&L (Class) (Individual)' in
Financials_Modul (output) 1.xlsx.

Column structure  (same as Excel):
  A  Particulars
  B  Mapping   (exact labels from Excel B-column)
  C  {company_name}          (= Company A)
  D  Concertiv Insurance Brokers, Inc.  (blank — no data in current upload)
  E  Consolidated             (= C + D)

Month-level logic (mirrors Excel B1 / C3 / D3 date-picker):
  • All P&L amounts are aggregated per month.
  • The preview ships ALL months; the frontend month-selector picks which
    month's C/D/E values to render — no extra API calls needed.
  • The downloaded Excel is written with ONE sheet per month.

SUMIFS formula equivalents (column-by-column):
  Revenue  →  SUMIFS(Amount, _Classification2=A[r], _Month=month)
  COGS     →  SUMIFS(Amount, _Classification3=A[r], _DeptClassOut="Cost of Revenue", _Month=month)
  OpEx     →  SUMIFS(Amount, _Classification2=A[r], _DeptClassOut=dept, _Month=month)
  D&A / Other / Tax  →  SUMIFS(Amount, _Classification2=A[r], _Month=month)

Calculated rows (identical to Excel formulas):
  Gross Profit       = Total Revenue − Total COGS            (C20 = C7−C15)
  Gross Profit (%)   = Gross Profit  / Client Recurring Rev  (C21 = C20/C8)
  Operating Profit   = Gross Profit  − Total OpEx            (C44 = C20−C24)
  Operating Profit % = Operating Profit / Total Revenue      (C45 = C44/C7)
  EBITDA             = Operating Profit + D&A                (C47 = C44+C43)
  EBITDA %           = EBITDA / Total Revenue                (C48 = C47/C7)
  Net Income         = Op Profit + Other − Tax               (C55 = C44+C50−C54)
  Net Income %       = Net Income / Total Revenue            (C56 = C55/C7)
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Row definitions — Mapping column values match the Excel B-column exactly ──

REVENUE_LINES = [
    # (A-col label,              B-col mapping label,             _Classification2 filter)
    ("Client Recurring",     "Revenue - Subscription",          "Client Recurring"),
    ("Billable Expense",     "Billable Expense Income",         "Billable Expense"),
    ("Project - One Time",   "Revenue - Project",               "Project - One Time"),
    ("Supplier Commission",  "Revenue - Supplier",              "Supplier Commission"),
    ("Intercompany Revenue", "Managed Services Income - I/C",   "Intercompany Revenue"),
    ("User Conference",      "Revenue - Conference",            "User Conference"),
]

COGS_LINES = [
    # (A-col label,              B-col mapping (="Cost of Revenue"), _Classification3 filter)
    ("Compensation",              "Cost of Revenue", "Compensation"),
    ("Technology",                "Cost of Revenue", "Technology"),
    ("Office Supplies/Equipment", "Cost of Revenue", "Office Supplies/Equipment"),
    ("Travel",                    "Cost of Revenue", "Travel"),
]

OPEX_DEPTS = [
    "Sales",
    "Marketing",
    "Customer Success",
    "Product",
    "R&D",
    "G&A",
]

# Sub-lines within each dept — B-col = department name (exact Excel B-column value)
OPEX_SUB_LINES = [
    ("Compensation", "Compensation"),
    ("Other",        "Other"),
]


# ── Sort months chronologically ───────────────────────────────────────────────

def _sort_months(months: list[str]) -> list[str]:
    def _key(m):
        try:
            return datetime.strptime(m, "%b-%y")
        except Exception:
            return datetime.min
    return sorted(months, key=_key)


# ── Aggregate P&L data by month ───────────────────────────────────────────────

def _aggregate(df: pd.DataFrame) -> tuple[dict, list]:
    """
    Returns agg[month][key] = float  (all data = Company A; Company B always 0).
    Key format:
      ("rev",  label)          – revenue lines
      ("cogs", label)          – cogs lines
      ("opex", dept, sub)      – opex lines
      ("da",)                  – depreciation & amortisation
      ("other_inc",)           – other income
      ("other_exp",)           – other expenses
      ("tax",)                 – tax expense
    """
    pl = df[df["_Financials"] == "Profit and Loss A/c"].copy()
    pl = pl[pl["_Month"].notna() & pl["Amount"].notna()]

    if pl.empty:
        return {}, []

    months = _sort_months(pl["_Month"].unique().tolist())
    agg: dict[str, dict] = {}

    for month in months:
        m = pl[pl["_Month"] == month]
        d: dict = {}

        # Revenue
        for label, _, cls2 in REVENUE_LINES:
            d[("rev", label)] = float(m.loc[m["_Classification2"] == cls2, "Amount"].sum())

        # COGS (Classification3 + DeptClassOut = "Cost of Revenue")
        for label, _, cls3 in COGS_LINES:
            mask = (m["_Classification3"] == cls3) & (m["_DeptClassOut"] == "Cost of Revenue")
            d[("cogs", label)] = float(m.loc[mask, "Amount"].sum())

        # OpEx by dept (Classification2 + DeptClassOut = dept)
        for dept in OPEX_DEPTS:
            for sub, cls2 in OPEX_SUB_LINES:
                mask = (m["_Classification2"] == cls2) & (m["_DeptClassOut"] == dept)
                d[("opex", dept, sub)] = float(m.loc[mask, "Amount"].sum())

        # D&A
        d[("da",)] = float(
            m.loc[m["_Classification2"] == "Depreciation and Amortization", "Amount"].sum()
        )
        # Other Income / Expenses
        d[("other_inc",)] = float(
            m.loc[m["_Classification2"] == "Other Income", "Amount"].sum()
        )
        d[("other_exp",)] = float(
            m.loc[m["_Classification2"] == "Other Expenses", "Amount"].sum()
        )
        # Tax
        d[("tax",)] = float(
            m.loc[m["_Classification2"] == "Tax Expense", "Amount"].sum()
        )
        agg[month] = d

    return agg, months


# ── Build structured row list ─────────────────────────────────────────────────

def _co(co_a: float, co_b: float = 0.0) -> dict:
    """Return {co_a, co_b, cons} with consolidated = co_a + co_b."""
    return {"co_a": co_a, "co_b": co_b, "cons": (co_a or 0.0) + (co_b or 0.0)}


def _build_rows(agg: dict, months: list) -> list[tuple]:
    """
    Returns list of (label, mapping, {month: {co_a, co_b, cons}} | None, row_type).

    Row types:
      section    – dark navy header (INCOME, COGS, OPERATING EXPENSES, …)
      subsection – lighter header (Revenue, Sales, Marketing, …)
      line       – individual data line
      subtotal   – per-dept total (Total Sales, Total G&A, …)
      total      – section total (Total Revenue, Total COGS, Total OpEx)
      grand_total– financial metric (GROSS PROFIT, OPERATING PROFIT, EBITDA, NET INCOME)
      metric     – percentage row (GP%, Op%, EBITDA%, NI%)
      blank      – spacer row
    """
    R: list[tuple] = []

    def va(key) -> dict:
        """Values for a raw aggregation key, company structure."""
        return {m: _co(agg[m].get(key, 0.0)) for m in months}

    def pct(num: dict, den: dict) -> dict:
        result = {}
        for m in months:
            d = den[m]["co_a"]
            n = num[m]["co_a"]
            result[m] = _co(
                (n / d * 100) if d and abs(d) > 0.001 else None,
            )
        return result

    def sum_vals(dicts: list[dict]) -> dict:
        result = {}
        for m in months:
            co_a = sum(d[m]["co_a"] for d in dicts)
            result[m] = _co(co_a)
        return result

    # ── INCOME ────────────────────────────────────────────────────────────────
    R.append(("Income", None, None, "section"))
    R.append(("Revenue", None, None, "subsection"))

    rev_dicts: list[dict] = []
    client_recurring: dict = {m: _co(0.0) for m in months}

    for label, mapping, _ in REVENUE_LINES:
        rv = va(("rev", label))
        if label == "Client Recurring":
            client_recurring = rv
        if any(abs(rv[m]["co_a"]) > 0.001 for m in months):
            R.append((f"   {label}", mapping, rv, "line"))
        rev_dicts.append(rv)

    total_rev = sum_vals(rev_dicts)
    R.append(("Total Revenue", None, total_rev, "total"))
    R.append((None, None, None, "blank"))

    # ── COGS ─────────────────────────────────────────────────────────────────
    R.append(("COGS", None, None, "section"))
    R.append(("Cost of Revenue", None, None, "subsection"))

    cogs_dicts: list[dict] = []
    for label, mapping, _ in COGS_LINES:
        cv = va(("cogs", label))
        if any(abs(cv[m]["co_a"]) > 0.001 for m in months):
            R.append((f"   {label}", mapping, cv, "line"))
        cogs_dicts.append(cv)

    total_cogs = sum_vals(cogs_dicts)
    R.append(("Total Cost of Revenue", None, total_cogs, "total"))
    R.append((None, None, None, "blank"))

    # ── GROSS PROFIT ──────────────────────────────────────────────────────────
    gross_profit = {m: _co(total_rev[m]["co_a"] - total_cogs[m]["co_a"]) for m in months}
    R.append(("Gross Profit", None, gross_profit, "grand_total"))
    # GP% = Gross Profit / Client Recurring Revenue  (mirrors C21 = C20/C8)
    R.append(("Gross Profit (%)", None, pct(gross_profit, client_recurring), "metric"))
    R.append((None, None, None, "blank"))

    # ── OPERATING EXPENSES ────────────────────────────────────────────────────
    R.append(("Expenses", None, None, "section"))
    R.append(("Operating Expenses", None, None, "subsection"))

    dept_total_dicts: list[dict] = []
    for dept in OPEX_DEPTS:
        R.append((f"   {dept}", None, None, "subsection"))
        sub_dicts: list[dict] = []
        for sub, _ in OPEX_SUB_LINES:
            sv = va(("opex", dept, sub))
            if any(abs(sv[m]["co_a"]) > 0.001 for m in months):
                R.append((f"   {sub}", dept, sv, "line"))   # B-col = dept name (matches Excel)
            sub_dicts.append(sv)
        dept_tot = sum_vals(sub_dicts)
        if any(abs(dept_tot[m]["co_a"]) > 0.001 for m in months):
            R.append((f"   Total {dept}", None, dept_tot, "subtotal"))
        dept_total_dicts.append(dept_tot)

    da_val = va(("da",))
    if any(abs(da_val[m]["co_a"]) > 0.001 for m in months):
        R.append(("   Depreciation and Amortization", None, da_val, "line"))
    dept_total_dicts.append(da_val)

    total_opex = sum_vals(dept_total_dicts)
    R.append(("   Total Operating Expenses", None, total_opex, "total"))
    R.append((None, None, None, "blank"))

    # ── OPERATING PROFIT ──────────────────────────────────────────────────────
    op_profit = {m: _co(gross_profit[m]["co_a"] - total_opex[m]["co_a"]) for m in months}
    R.append(("Operating Profit", None, op_profit, "grand_total"))
    # Op% = Operating Profit / Total Revenue  (mirrors C45 = C44/C7)
    R.append(("Operating Profit (%)", None, pct(op_profit, total_rev), "metric"))
    R.append((None, None, None, "blank"))

    # ── EBITDA ────────────────────────────────────────────────────────────────
    ebitda = {m: _co(op_profit[m]["co_a"] + da_val[m]["co_a"]) for m in months}
    R.append(("EBITDA", None, ebitda, "grand_total"))
    # EBITDA% = EBITDA / Total Revenue  (mirrors C48 = C47/C7)
    R.append(("EBITDA (%)", None, pct(ebitda, total_rev), "metric"))
    R.append((None, None, None, "blank"))

    # ── OTHER INCOME (EXPENSES) ───────────────────────────────────────────────
    R.append(("Other Income (Expenses)", None, None, "section"))

    oi = va(("other_inc",))
    oe = va(("other_exp",))
    if any(abs(oi[m]["co_a"]) > 0.001 for m in months):
        R.append(("   Other Income", None, oi, "line"))
    if any(abs(oe[m]["co_a"]) > 0.001 for m in months):
        R.append(("   Other Expenses", None, oe, "line"))

    total_other = {m: _co(oi[m]["co_a"] + oe[m]["co_a"]) for m in months}
    R.append(("   Total Other Income (Expenses)", None, total_other, "total"))
    R.append((None, None, None, "blank"))

    # ── TAX & NET INCOME ──────────────────────────────────────────────────────
    R.append(("Tax Expense", None, None, "section"))
    tax_val = va(("tax",))
    if any(abs(tax_val[m]["co_a"]) > 0.001 for m in months):
        R.append(("   Tax Expense", None, tax_val, "line"))

    # Net Income = Op Profit + Other − Tax  (mirrors C55 = C44+C50−C54)
    net_income = {m: _co(op_profit[m]["co_a"] + total_other[m]["co_a"] - tax_val[m]["co_a"]) for m in months}
    R.append(("Net Income", None, net_income, "grand_total"))
    # NI% = Net Income / Total Revenue  (mirrors C56 = C55/C7)
    R.append(("Net Income (%)", None, pct(net_income, total_rev), "metric"))

    return R


# ── Excel writer ──────────────────────────────────────────────────────────────

def run_pl_individual(df: pd.DataFrame, company_name: str) -> bytes:
    """Write Excel with the LATEST month, 5 columns (Particulars, Mapping, Co-A, Co-B, Cons)."""
    agg, months = _aggregate(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Base P&L (Class) (Individual)"

    if not months:
        ws["A1"] = "No P&L data found in the uploaded file."
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    rows = _build_rows(agg, months)
    latest = months[-1]

    # ── Styles ────────────────────────────────────────────────────────────────
    def F(bold=False, size=9, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    DARK_NAVY  = PatternFill("solid", fgColor="0F172A")
    DARK_BLUE  = PatternFill("solid", fgColor="1E3A5F")
    MED_BLUE   = PatternFill("solid", fgColor="EBF2FB")
    LIGHT_GRAY = PatternFill("solid", fgColor="F1F5F9")
    SUBSEC_BG  = PatternFill("solid", fgColor="F8FAFC")
    METRIC_BG  = PatternFill("solid", fgColor="F0F9FF")

    AL = Alignment(horizontal="left",   vertical="center")
    AR = Alignment(horizontal="right",  vertical="center")
    AC = Alignment(horizontal="center", vertical="center")

    thin  = Side(border_style="thin",   color="CBD5E1")
    thick = Side(border_style="medium", color="64748B")
    NUM   = "#,##0.00"
    PCTF  = '0.00%'

    # ── Title block ───────────────────────────────────────────────────────────
    ws.append([company_name, None, None, None, None])
    ws["A1"].font = F(bold=True, size=14, color="0F172A"); ws.row_dimensions[1].height = 22

    ws.append(["Base P&L (Class) (Individual)", None, None, None, None])
    ws["A2"].font = F(bold=True, size=11, color="1E3A5F")

    ws.append([f"Month: {latest}", None, None, None, None])
    ws["A3"].font = F(size=9, color="64748B", italic=True)

    ws.append([None] * 5)  # spacer

    # ── Column header row ─────────────────────────────────────────────────────
    headers = ["Particulars", "Mapping", company_name,
               "Concertiv Insurance Brokers, Inc.", "Consolidated"]
    ws.append(headers)
    hdr_r = ws.max_row
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(hdr_r, ci)
        c.font      = F(bold=True, size=9, color="FFFFFF")
        c.fill      = DARK_BLUE
        c.alignment = AL if ci == 1 else (AL if ci == 2 else AC)
    ws.row_dimensions[hdr_r].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    for label, mapping, values, rtype in rows:
        if rtype == "blank":
            ws.append([None] * 5); continue

        if values is None:
            co_a = co_b = cons = None
        else:
            cv = values.get(latest, {"co_a": 0.0, "co_b": 0.0, "cons": 0.0})
            co_a = cv["co_a"]
            co_b = cv["co_b"] if cv["co_b"] != 0.0 else None  # show blank for zero
            cons = cv["cons"]

        ws.append([label, mapping or "", co_a, co_b, cons])
        r = ws.max_row
        ca = ws.cell(r, 1); cb = ws.cell(r, 2)
        ca.alignment = AL; cb.alignment = AL

        for ci in (3, 4, 5):
            cell = ws.cell(r, ci)
            if cell.value is not None:
                if rtype == "metric":
                    cell.number_format = PCTF
                    cell.value = cell.value / 100 if cell.value else cell.value
                else:
                    cell.number_format = NUM
                cell.alignment = AR

        def fill_row(fill):
            for ci in range(1, 6): ws.cell(r, ci).fill = fill

        if rtype == "section":
            fill_row(DARK_NAVY)
            ca.font = F(bold=True, size=9, color="FFFFFF")
            cb.font = F(bold=True, size=9, color="FFFFFF")
            for ci in (3, 4, 5): ws.cell(r, ci).font = F(bold=True, size=9, color="FFFFFF")
            ws.row_dimensions[r].height = 15
        elif rtype == "subsection":
            fill_row(SUBSEC_BG)
            ca.font = F(bold=True, size=9, color="1E293B")
        elif rtype == "line":
            ca.font = F(size=9, color="334155")
            cb.font = F(size=9, color="64748B", italic=True)
            for ci in (3, 4, 5):
                cell = ws.cell(r, ci)
                if cell.value is not None:
                    cell.font = F(size=9, color="DC2626" if cell.value < 0 else "334155")
        elif rtype == "subtotal":
            fill_row(MED_BLUE)
            for ci in range(1, 6):
                c = ws.cell(r, ci)
                c.font = F(bold=True, size=9, color="1E40AF")
                c.border = Border(top=thin)
        elif rtype == "total":
            fill_row(LIGHT_GRAY)
            for ci in range(1, 6):
                c = ws.cell(r, ci)
                c.font = F(bold=True, size=9)
                c.border = Border(top=thin)
        elif rtype == "grand_total":
            fill_row(DARK_BLUE)
            for ci in range(1, 6):
                c = ws.cell(r, ci)
                val = c.value
                c.font = F(bold=True, size=10, color="FCA5A5" if isinstance(val, (int, float)) and val < 0 else "FFFFFF")
                c.border = Border(top=Side(border_style="medium", color="64748B"))
            ws.row_dimensions[r].height = 16
        elif rtype == "metric":
            fill_row(METRIC_BG)
            ca.font = F(italic=True, size=8, color="0369A1")
            cb.font = F(italic=True, size=8, color="64748B")
            for ci in (3, 4, 5):
                c = ws.cell(r, ci)
                if c.value is not None:
                    c.font = F(italic=True, size=8, color="DC2626" if c.value < 0 else "0369A1")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 15
    ws.freeze_panes = "C6"

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── Preview payload ───────────────────────────────────────────────────────────

def get_pl_individual_preview(df: pd.DataFrame, company_name: str = "") -> dict:
    """
    Return JSON-serialisable preview for the frontend.

    Shape:
      {
        months: ["Jan-24", ...],
        company_name: "...",
        rows: [
          {
            label:   str | null,
            mapping: str | null,
            values:  { month: {co_a, co_b, cons} } | null,
            type:    str
          }
        ]
      }
    """
    agg, months = _aggregate(df)
    if not months:
        return {"months": [], "company_name": company_name, "rows": []}

    rows = _build_rows(agg, months)

    def _round_cv(cv):
        if cv is None:
            return None
        return {
            "co_a": round(cv["co_a"], 2) if cv["co_a"] is not None else None,
            "co_b": round(cv["co_b"], 2) if cv["co_b"] is not None else None,
            "cons": round(cv["cons"], 2) if cv["cons"] is not None else None,
        }

    serialisable = []
    for label, mapping, values, rtype in rows:
        if values is not None:
            vals = {m: _round_cv(v) for m, v in values.items()}
        else:
            vals = None
        serialisable.append({
            "label":   label,
            "mapping": mapping,
            "values":  vals,
            "type":    rtype,
        })

    return {
        "months":       months,
        "company_name": company_name,
        "rows":         serialisable,
    }
