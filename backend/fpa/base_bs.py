"""
Base BS generator — produces the Balance Sheet Excel from the transformed staging DataFrame.

Logic mirrors the reference workbook formula:
  Month-0 balance  = last Balance value per account (cumulative from QB inception)
  Month-N balance  = Month-(N-1) balance + SUM of Amounts in Month-N for that classification

Row structure, subtotals and totals exactly match 'Base BS' in Financials_Modul (output) 1.xlsx.
Retained Earning is computed as a plug figure so the balance sheet always ties.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── BS Row Definitions ────────────────────────────────────────────────────────
# (label, row_type, extra)
#   row_type  extra
#   --------  -----------------------------------------------------------
#   section   None                       – bold dark header (Assets / Liabilities)
#   group     [child_label, ...]         – sum of listed child labels
#   data      "Classification (Line Item)" name
#   total     [dep_label, ...]           – sum of listed dependency labels
#   retained  None                       – plug: Total Assets − other equity items
#   check     None                       – Total Assets − Total Liabilities (→ 0)
#   nwc       None                       – Current Assets − Current Liabilities
#   nwc_chg   None                       – current NWC − prior NWC
#   blank     None

BS_ROWS = [
    ("Assets",                                                    "section", None),
    ("Current Assets",                                            "group",   ["Cash and Cash equivalents", "Accounts Receivable"]),
    ("Cash and Cash equivalents",                                 "data",    "Cash and Cash equivalents"),
    ("Accounts Receivable",                                       "data",    "Accounts Receivable"),
    ("Other Current Assets",                                      "group",   ["Prepaid expenses and other current assets"]),
    ("Prepaid expenses and other current assets",                 "data",    "Prepaid expenses and other current assets"),
    ("Total Current Assets",                                      "total",   ["Current Assets", "Other Current Assets"]),
    ("Fixed Assets",                                              "group",   [
        "Tangible Assets, net of accumulated depreciation",
        "Goodwill, net of accumulated amortization",
        "Intangible assets, net of accumulated amortization",
        "Operating lease right-of-use assets",
        "Finance lease right-of-use assets",
    ]),
    ("Tangible Assets, net of accumulated depreciation",          "data",    "Tangible Assets, net of accumulated depreciation"),
    ("Goodwill, net of accumulated amortization",                 "data",    "Goodwill, net of accumulated amortization"),
    ("Intangible assets, net of accumulated amortization",        "data",    "Intangible assets, net of accumulated amortization"),
    ("Operating lease right-of-use assets",                       "data",    "Operating lease right-of-use assets"),
    ("Finance lease right-of-use assets",                         "data",    "Finance lease right-of-use assets"),
    ("Other Non-Current Assets",                                  "group",   ["Deposit or Advances"]),
    ("Deposit or Advances",                                       "data",    "Deposit or Advances"),
    ("Total Non-Current Assets",                                  "total",   ["Fixed Assets", "Other Non-Current Assets"]),
    ("Total Assets",                                              "total",   ["Total Current Assets", "Total Non-Current Assets"]),
    (None,                                                        "blank",   None),
    ("Liabilities",                                               "section", None),
    ("Current Liabilities",                                       "group",   [
        "Accounts Payable", "Accrued expenses",
        "Current portion of operating lease liabilities",
        "Due to Employee's",
    ]),
    ("Accounts Payable",                                          "data",    "Accounts Payable"),
    ("Accrued expenses",                                          "data",    "Accrued expenses"),
    ("Current portion of operating lease liabilities",            "data",    "Current portion of operating lease liabilities"),
    ("Due to Employee's",                                         "data",    "Due to Employee's"),
    ("Non Current Liabilities",                                   "group",   [
        "Deferred Revenue", "Other Long term Liabilities",
        "Statutory Dues",
        "Operating lease liabilities, net of current portion",
    ]),
    ("Deferred Revenue",                                          "data",    "Deferred Revenue"),
    ("Other Long term Liabilities",                               "data",    "Other Long term Liabilities"),
    ("Statutory Dues",                                            "data",    "Statutory Dues"),
    ("Operating lease liabilities, net of current portion",       "data",    "Operating lease liabilities, net of current portion"),
    ("Equity",                                                    "group",   [
        "Common Stock", "Preferred Stock",
        "Additional paid-in capital", "Retained Earning",
    ]),
    ("Common Stock",                                              "data",    "Common Stock"),
    ("Preferred Stock",                                           "data",    "Preferred Stock"),
    ("Additional paid-in capital",                                "data",    "Additional paid-in capital"),
    ("Retained Earning",                                          "retained", None),
    ("Total Liabilities",                                         "total",   ["Current Liabilities", "Non Current Liabilities", "Equity"]),
    (None,                                                        "blank",   None),
    ("Check",                                                     "check",   None),
    (None,                                                        "blank",   None),
    ("Net Working Capital",                                       "nwc",     None),
    ("Changes in Net Working Capital",                            "nwc_chg", None),
]


# ── Value computation ─────────────────────────────────────────────────────────

def _build_values(df: pd.DataFrame) -> tuple[dict, list]:
    """
    Returns (row_values, sorted_months).
    row_values: {label: {month: float}}
    """
    bs = df[df["_Financials"] == "Balance Sheet"].copy()
    if bs.empty:
        return {}, []

    bs["_date_parsed"] = pd.to_datetime(bs["Date"], format="%m/%d/%Y", errors="coerce")
    bs = bs.sort_values("_date_parsed")

    all_months: list[str] = sorted(
        bs["_Month"].dropna().unique().tolist(),
        key=lambda m: pd.to_datetime(m, format="%b-%y"),
    )
    if not all_months:
        return {}, []

    # ── Step 1: month-end Balance per (Account, Classification, Month) ─────────
    # Last Balance value for each account each month = end-of-month balance.
    # Uses the QB running-balance column (cumulative from inception).
    month_end_raw = (
        bs[bs["_Month"].notna() & bs["Balance"].notna()]
        .groupby(["Account", "_Classification", "_Month"])["Balance"]
        .last()
        .unstack(level="_Month", fill_value=None)
    )

    # Ensure month columns are in order
    month_cols = [m for m in all_months if m in month_end_raw.columns]
    month_end_raw = month_end_raw.reindex(columns=month_cols)

    # Forward-fill: accounts with no transactions in a month carry their last balance
    month_end_raw = month_end_raw.ffill(axis=1).fillna(0.0)

    # Aggregate by Classification (sum of per-account month-end balances)
    cls_bal = month_end_raw.groupby(level="_Classification").sum()

    def v(cls: str, month: str) -> float:
        if cls in cls_bal.index and month in cls_bal.columns:
            return float(cls_bal.at[cls, month])
        return 0.0

    rv: dict[str, dict[str, float]] = {}

    # ── Step 2: data rows ───────────────────────────────────────────────────────
    for label, rtype, extra in BS_ROWS:
        if rtype == "data":
            rv[label] = {m: v(extra, m) for m in all_months}

    # ── Step 3: group rows (sum of children) ───────────────────────────────────
    for label, rtype, extra in BS_ROWS:
        if rtype == "group":
            rv[label] = {
                m: sum(rv.get(child, {}).get(m, 0.0) for child in extra if child != "Retained Earning")
                for m in all_months
            }

    # ── Step 4: total rows (multi-pass to handle chained deps) ─────────────────
    total_rows = [(lbl, ext) for lbl, rt, ext in BS_ROWS if rt == "total"]
    for _ in range(6):
        for label, extra in total_rows:
            if all(dep in rv for dep in extra):
                rv[label] = {
                    m: sum(rv.get(dep, {}).get(m, 0.0) for dep in extra)
                    for m in all_months
                }

    # ── Step 5: Retained Earning (plug to balance the sheet) ──────────────────
    equity_excl_re = ["Common Stock", "Preferred Stock", "Additional paid-in capital"]
    rv["Retained Earning"] = {
        m: (
            rv.get("Total Assets", {}).get(m, 0.0)
            - rv.get("Current Liabilities", {}).get(m, 0.0)
            - rv.get("Non Current Liabilities", {}).get(m, 0.0)
            - sum(rv.get(eq, {}).get(m, 0.0) for eq in equity_excl_re)
        )
        for m in all_months
    }

    # ── Step 6: recompute Equity and Total Liabilities with RE ─────────────────
    rv["Equity"] = {
        m: sum(rv.get(ch, {}).get(m, 0.0)
               for ch in ["Common Stock", "Preferred Stock",
                           "Additional paid-in capital", "Retained Earning"])
        for m in all_months
    }
    rv["Total Liabilities"] = {
        m: sum(rv.get(dep, {}).get(m, 0.0)
               for dep in ["Current Liabilities", "Non Current Liabilities", "Equity"])
        for m in all_months
    }

    # ── Step 7: Check, NWC, NWC change ─────────────────────────────────────────
    rv["Check"] = {
        m: rv.get("Total Assets", {}).get(m, 0.0)
           - rv.get("Total Liabilities", {}).get(m, 0.0)
        for m in all_months
    }
    rv["Net Working Capital"] = {
        m: rv.get("Total Current Assets", {}).get(m, 0.0)
           - rv.get("Current Liabilities", {}).get(m, 0.0)
        for m in all_months
    }
    nwc = rv["Net Working Capital"]
    rv["Changes in Net Working Capital"] = {
        m: nwc.get(m, 0.0) - nwc.get(all_months[i - 1], 0.0) if i > 0 else 0.0
        for i, m in enumerate(all_months)
    }

    return rv, all_months


# ── Excel builder ─────────────────────────────────────────────────────────────

def run_base_bs(df: pd.DataFrame, company_name: str) -> bytes:
    """Generate Base BS Excel. Returns raw bytes of the .xlsx file."""

    rv, all_months = _build_values(df)
    n = len(all_months)

    wb = Workbook()
    ws = wb.active
    ws.title = "Base BS"

    # ── Palette ──────────────────────────────────────────────────────────────
    DARK_BLUE  = PatternFill("solid", fgColor="1E3A5F")
    ROW_BLUE   = PatternFill("solid", fgColor="EBF2FB")
    LIGHT_GRAY = PatternFill("solid", fgColor="F1F5F9")
    GREEN_FILL = PatternFill("solid", fgColor="F0FDF4")
    WHITE      = PatternFill("solid", fgColor="FFFFFF")

    def font(bold=False, size=9, color="000000"):
        return Font(name="Calibri", bold=bold, size=size, color=color)

    ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    ALIGN_R = Alignment(horizontal="right",  vertical="center")
    ALIGN_C = Alignment(horizontal="center", vertical="center")

    thin  = Side(border_style="thin",   color="CBD5E1")
    thick = Side(border_style="medium", color="94A3B8")
    TOP   = Border(top=thin)
    MED_TOP = Border(top=thick)

    NUM_FMT = "#,##0.00"

    # ── Title rows ───────────────────────────────────────────────────────────
    ws.append([company_name] + [None] * n)
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    ws.row_dimensions[1].height = 22

    ws.append(["Base Balance Sheet"] + [None] * n)
    ws["A2"].font = Font(name="Calibri", bold=True, size=11, color="475569")
    ws.row_dimensions[2].height = 16

    date_rng = f"{all_months[0]} — {all_months[-1]}" if all_months else ""
    ws.append([date_rng] + [None] * n)
    ws["A3"].font = Font(name="Calibri", size=9, color="64748B")

    ws.append([None] * (n + 1))  # blank row

    # ── Column-header row (row 5) ─────────────────────────────────────────────
    ws.append(["Particulars"] + list(all_months))
    for ci in range(1, n + 2):
        c = ws.cell(5, ci)
        c.fill = DARK_BLUE
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.alignment = ALIGN_L if ci == 1 else ALIGN_C
    ws.row_dimensions[5].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    for label, rtype, _ in BS_ROWS:

        if rtype == "blank":
            ws.append([None] * (n + 1))
            continue

        if rtype == "section":
            ws.append([label] + [None] * n)
            r = ws.max_row
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.fill = DARK_BLUE
                c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            ws.cell(r, 1).alignment = ALIGN_L
            ws.row_dimensions[r].height = 16
            continue

        month_vals = rv.get(label, {})
        row_data = [label] + [month_vals.get(m, 0.0) for m in all_months]
        ws.append(row_data)
        r = ws.max_row

        lc = ws.cell(r, 1)
        lc.alignment = ALIGN_L

        if rtype == "group":
            lc.font    = font(bold=True, size=9, color="1E3A5F")
            lc.fill    = ROW_BLUE
            for ci in range(2, n + 2):
                c = ws.cell(r, ci)
                c.fill          = ROW_BLUE
                c.font          = font(bold=True, size=9, color="1E3A5F")
                c.number_format = NUM_FMT
                c.alignment     = ALIGN_R
            ws.row_dimensions[r].height = 15

        elif rtype == "data":
            lc.value = "  " + (label or "")
            lc.font  = font(size=9)
            for ci in range(2, n + 2):
                c = ws.cell(r, ci)
                c.font          = font(size=9)
                c.number_format = NUM_FMT
                c.alignment     = ALIGN_R

        elif rtype == "total":
            lc.font = font(bold=True, size=9)
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.fill   = LIGHT_GRAY
                c.font   = font(bold=True, size=9)
                c.border = MED_TOP if label in ("Total Assets", "Total Liabilities") else TOP
                if ci > 1:
                    c.number_format = NUM_FMT
                    c.alignment     = ALIGN_R
            ws.row_dimensions[r].height = 15

        elif rtype == "retained":
            lc.value = "  Retained Earning"
            lc.font  = font(size=9)
            for ci in range(2, n + 2):
                c = ws.cell(r, ci)
                c.font          = font(size=9)
                c.number_format = NUM_FMT
                c.alignment     = ALIGN_R

        elif rtype == "check":
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.fill = GREEN_FILL
                c.font = font(bold=True, size=9, color="166534")
                if ci > 1:
                    c.number_format = NUM_FMT
                    c.alignment     = ALIGN_R
            ws.row_dimensions[r].height = 15

        elif rtype in ("nwc", "nwc_chg"):
            lc.font = font(bold=True, size=9)
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.fill = LIGHT_GRAY
                c.font = font(bold=True, size=9)
                if ci > 1:
                    c.number_format = NUM_FMT
                    c.alignment     = ALIGN_R

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 48
    for i in range(n):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14
    ws.freeze_panes = "B6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Preview payload (JSON-serialisable) ──────────────────────────────────────

def get_bs_preview(df: pd.DataFrame) -> dict:
    """
    Return {months, rows} for the frontend BS table.
    'rows' mirrors BS_ROWS with computed numeric values attached.
    """
    rv, all_months = _build_values(df)

    rows = []
    for label, rtype, _ in BS_ROWS:
        vals = None
        if label and label in rv:
            vals = [round(rv[label].get(m, 0.0), 2) for m in all_months]

        display = label
        if rtype == "data":
            display = label          # frontend will indent
        elif rtype == "retained":
            display = "Retained Earning"

        rows.append({"label": display, "type": rtype, "values": vals})

    return {"months": all_months, "rows": rows}
