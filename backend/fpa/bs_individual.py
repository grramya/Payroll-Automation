"""
BS (Individual) generator — mirrors the layout and logic of
'BS (Individual)' in Financials_Modul (output) 1.xlsx.

Column structure (same as Excel, rows 3-5):
  A  Particulars
  B  {company_name}                       (= Concertiv Inc.)
  C  Concertiv Insurance Brokers, Inc.    (blank — no data in current upload)
  D  Eliminations                         (blank — no inter-company data)
  E  Consolidated BS                      (= B + C + D)

Month-selector logic:
  • The Excel shows balances "As of {date}" — changing the date recalculates.
  • We replicate this by computing, for each available BS month, the last
    Balance per account using only rows on or before that month.
  • Preview ships ALL months; the frontend month-picker switches views.

Value logic:
  1. For a given "as-of" month, filter BS rows to months ≤ as-of month.
  2. Sort by date; take the LAST Balance per account (cumulative running
     balance — the last entry IS the end-of-period account balance).
  3. Net Income = cumulative sum of P&L Amounts through the as-of month.
  4. Retained Earnings = plug: Total Assets − Liabilities − CS − PS − APIC − NI.
"""

import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _sort_months(months: list) -> list:
    """Sort 'Mmm-yy' strings chronologically."""
    def _key(m):
        try:
            return datetime.strptime(m, "%b-%y")
        except Exception:
            return datetime.min
    return sorted(months, key=_key)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _leaf(full_name: str) -> str:
    """Strip leading account number; return leaf segment of colon-separated name.
    '131010 Prepaid Expenses:Prepaid Expense' → 'Prepaid Expense'
    '120010 Accounts Receivable' → 'Accounts Receivable'
    """
    n = re.sub(r"^\d+\s+", "", str(full_name).strip())
    return n.split(":")[-1].strip() if ":" in n else n


# ── Intercompany account detection ───────────────────────────────────────────

_IC_KEYWORDS = ("intercompany receivable", "intercompany payable")


def _is_ic_account(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in _IC_KEYWORDS)


# ── Classification → BS section mapping ──────────────────────────────────────

# Maps each BS (Individual) section group to a list of Classification (Line Item) values
BS_GROUPS: dict[str, list[str]] = {
    "Bank Accounts":         ["Cash and Cash equivalents"],
    "Accounts Receivable":   ["Accounts Receivable"],
    "Other Current Assets":  ["Prepaid expenses and other current assets",
                               "Deposit or Advances"],
    "Fixed Assets":          ["Tangible Assets, net of accumulated depreciation"],
    "Goodwill":              ["Goodwill, net of accumulated amortization"],
    "Intangible Assets":     ["Intangible assets, net of accumulated amortization"],
    "ROU Operating Leases":  ["Operating lease right-of-use assets"],
    "ROU Finance Leases":    ["Finance lease right-of-use assets"],
    "Accounts Payable":      ["Accounts Payable"],
    "Accrued Expenses":      ["Accrued expenses"],
    "Deferred Revenue":      ["Deferred Revenue"],
    "Due to Employees":      ["Due to Employee's"],
    "Lease - Current":       ["Current portion of operating lease liabilities"],
    "Statutory Dues":        ["Statutory Dues"],
    "Long-Term Liabilities": ["Operating lease liabilities, net of current portion",
                               "Other Long term Liabilities"],
    "Common Stock":          ["Common Stock"],
    "Preferred Stock":       ["Preferred Stock"],
    "APIC":                  ["Additional paid-in capital"],
}


# ── Value computation ─────────────────────────────────────────────────────────

def _compute(df: pd.DataFrame, as_of_month: str | None = None) -> tuple[dict, float, str]:
    """
    Returns:
      acct_balances  – {group_key: {account_display_name: balance}}
      net_income     – cumulative sum of P&L Amounts through as_of_month
      as_of_label    – e.g. 'Apr-26'

    Operates on all rows in df regardless of _source.  Use _compute_split for
    per-company results when df has a '_source' column.
    If as_of_month is given, only BS/P&L rows up to and including that month
    are considered, matching the Excel "As of {date}" date-picker logic.
    """
    df = df.copy()
    df["_date_parsed"] = pd.to_datetime(df["Date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("_date_parsed")

    bs_all = df[df["_Financials"] == "Balance Sheet"].copy()
    bs_all = bs_all[bs_all["_Month"].notna() & bs_all["Balance"].notna()]
    pl_all = df[df["_Financials"] == "Profit and Loss A/c"].copy()

    all_bs_months = _sort_months(bs_all["_Month"].dropna().unique().tolist())
    all_pl_months = _sort_months(pl_all["_Month"].dropna().unique().tolist())

    if as_of_month and as_of_month in all_bs_months:
        idx = all_bs_months.index(as_of_month)
        bs = bs_all[bs_all["_Month"].isin(set(all_bs_months[:idx + 1]))]
        as_of_label = as_of_month
    else:
        bs = bs_all
        as_of_label = all_bs_months[-1] if all_bs_months else "N/A"

    if as_of_month and as_of_month in all_pl_months:
        idx = all_pl_months.index(as_of_month)
        pl = pl_all[pl_all["_Month"].isin(set(all_pl_months[:idx + 1]))]
    else:
        pl = pl_all

    net_income = float(pl["Amount"].sum())

    # Last Balance per account within the filtered window
    last_bal = (
        bs.sort_values("_date_parsed")
        .groupby(["Account", "_Classification"])["Balance"]
        .last()
    )

    acct_balances: dict[str, dict[str, float]] = {g: {} for g in BS_GROUPS}
    for (acct, cls), bal in last_bal.items():
        for group, cls_list in BS_GROUPS.items():
            if cls in cls_list:
                display = _leaf(acct)
                key = display
                suffix = 1
                while key in acct_balances[group]:
                    key = f"{display} ({suffix})"
                    suffix += 1
                acct_balances[group][key] = float(bal)
                break

    return acct_balances, net_income, as_of_label


def _compute_split(
    df: pd.DataFrame, as_of_month: str | None = None
) -> tuple[dict, float, dict, float, str]:
    """
    When df has a '_source' column, computes balance sheet separately for
    'main' and 'broker' companies.
    Returns (acct_a, net_a, acct_b, net_b, as_of_label).
    Falls back to treating all rows as co_a if '_source' is absent.
    """
    if "_source" in df.columns:
        df_a = df[df["_source"] == "main"]
        df_b = df[df["_source"] == "broker"]
    else:
        df_a = df
        df_b = df.iloc[0:0]

    acct_a, net_a, as_of = _compute(df_a, as_of_month)
    if df_b.empty:
        empty_b = {g: {} for g in acct_a}
        return acct_a, net_a, empty_b, 0.0, as_of

    acct_b, net_b, _ = _compute(df_b, as_of_month)
    return acct_a, net_a, acct_b, net_b, as_of


def _group_total(acct_balances: dict, groups: list[str]) -> float:
    return sum(
        bal
        for g in groups
        for bal in acct_balances.get(g, {}).values()
    )


# ── Build row list ────────────────────────────────────────────────────────────

def _build_rows(
    acct_balances: dict,
    net_income: float,
    acct_balances_b: dict | None = None,
    net_income_b: float = 0.0,
) -> list[tuple]:
    """
    Returns list of (label, co_a_value, co_b_value, elim_value, row_type).
    co_b_value and elim_value are None when acct_balances_b is not provided.
    elim_value is non-zero only for intercompany accounts (IC Receivable / Payable).
    row_type is one of: 'section' | 'subsection' | 'group_header' | 'account' |
      'group_total' | 'total' | 'grand_total' | 'equity_item' | 'check' | 'blank'
    """
    R = []
    has_b = acct_balances_b is not None

    def G_a(key):
        return acct_balances.get(key, {})

    def G_b(key):
        return (acct_balances_b or {}).get(key, {})

    def add_group(header_label, group_keys, indent="   "):
        """Append header + accounts + total. Returns (total_a, total_b, total_elim)."""
        if isinstance(group_keys, str):
            group_keys = [group_keys]
        total_a = 0.0
        total_b = 0.0
        total_elim = 0.0
        R.append((header_label, None, None, None, "group_header"))
        for gk in group_keys:
            all_names = sorted(set(G_a(gk)) | set(G_b(gk)))
            for name in all_names:
                val_a = G_a(gk).get(name)
                val_b = G_b(gk).get(name) if has_b else None
                elim = -((val_a or 0.0) + (val_b or 0.0)) if _is_ic_account(name) else 0.0
                total_a += (val_a or 0.0)
                total_b += (val_b or 0.0)
                total_elim += elim
                R.append((indent + "   " + name, val_a, val_b, elim or None, "account"))
        R.append(("Total " + header_label.strip(), total_a, total_b if has_b else None, total_elim or None, "group_total"))
        return total_a, total_b, total_elim

    def row(label, val_a, val_b, elim, rtype):
        R.append((label, val_a, val_b if has_b else None, elim, rtype))

    # ── ASSETS ───────────────────────────────────────────────────────────────
    row("ASSETS", None, None, None, "section")
    row("   Current Assets", None, None, None, "subsection")

    bank_a, bank_b, bank_e = add_group("      Bank Accounts",       "Bank Accounts")
    ar_a,   ar_b,   ar_e   = add_group("      Accounts Receivable", "Accounts Receivable")
    oca_a,  oca_b,  oca_e  = add_group("      Other Current Assets", ["Other Current Assets"], indent="      ")
    total_ca_a = bank_a + ar_a + oca_a
    total_ca_b = bank_b + ar_b + oca_b
    total_ca_e = bank_e + ar_e + oca_e
    row("   Total Current Assets", total_ca_a, total_ca_b, total_ca_e or None, "total")

    row("   Fixed Assets", None, None, None, "subsection")
    fa_a, fa_b, fa_e = add_group("      Fixed Assets", "Fixed Assets")
    row("   Total Fixed Assets", fa_a, fa_b, fa_e or None, "total")

    row("   Other Assets", None, None, None, "subsection")
    gw_a,     gw_b,     gw_e     = add_group("      Goodwill",               "Goodwill")
    ia_a,     ia_b,     ia_e     = add_group("      Intangible Assets",      "Intangible Assets")
    rou_op_a, rou_op_b, rou_op_e = add_group("      ROU - Operating Leases", "ROU Operating Leases")
    rou_fi_a, rou_fi_b, rou_fi_e = add_group("      ROU - Finance Leases",   "ROU Finance Leases")
    total_oa_a = gw_a + ia_a + rou_op_a + rou_fi_a
    total_oa_b = gw_b + ia_b + rou_op_b + rou_fi_b
    total_oa_e = gw_e + ia_e + rou_op_e + rou_fi_e
    row("   Total Other Assets", total_oa_a, total_oa_b, total_oa_e or None, "total")

    total_assets_a = total_ca_a + fa_a + total_oa_a
    total_assets_b = total_ca_b + fa_b + total_oa_b
    total_assets_e = total_ca_e + fa_e + total_oa_e
    row("TOTAL ASSETS", total_assets_a, total_assets_b, total_assets_e or None, "grand_total")

    row(None, None, None, None, "blank")

    # ── LIABILITIES AND EQUITY ────────────────────────────────────────────────
    row("LIABILITIES AND EQUITY", None, None, None, "section")
    row("   Liabilities", None, None, None, "subsection")
    row("      Current Liabilities", None, None, None, "subsection")

    ap_a,  ap_b,  ap_e  = add_group("         Accounts Payable",           "Accounts Payable",   indent="         ")
    row("         Other Current Liabilities", None, None, None, "subsection")
    ae_a,  ae_b,  ae_e  = add_group("            Accrued Expenses",        "Accrued Expenses",   indent="            ")
    dr_a,  dr_b,  dr_e  = add_group("            Deferred Revenue",        "Deferred Revenue",   indent="            ")
    dte_a, dte_b, dte_e = add_group("            Due to Employees",        "Due to Employees",   indent="            ")
    lcl_a, lcl_b, lcl_e = add_group("            Lease Payable - Current", "Lease - Current",    indent="            ")
    std_a, std_b, std_e = add_group("            Statutory Dues",          "Statutory Dues",     indent="            ")

    total_ocl_a = ae_a + dr_a + dte_a + lcl_a + std_a
    total_ocl_b = ae_b + dr_b + dte_b + lcl_b + std_b
    total_ocl_e = ae_e + dr_e + dte_e + lcl_e + std_e
    row("         Total Other Current Liabilities", total_ocl_a, total_ocl_b, total_ocl_e or None, "total")
    total_cl_a = ap_a + total_ocl_a
    total_cl_b = ap_b + total_ocl_b
    total_cl_e = ap_e + total_ocl_e
    row("      Total Current Liabilities", total_cl_a, total_cl_b, total_cl_e or None, "total")

    lt_a, lt_b, lt_e = add_group("      Long-Term Liabilities", "Long-Term Liabilities", indent="      ")
    total_liab_a = total_cl_a + lt_a
    total_liab_b = total_cl_b + lt_b
    total_liab_e = total_cl_e + lt_e
    row("   Total Liabilities", total_liab_a, total_liab_b, total_liab_e or None, "total")

    # ── Equity ────────────────────────────────────────────────────────────────
    row("   Equity", None, None, None, "subsection")
    cs_a,   cs_b,   _cs_e   = add_group("      Common Stock",               "Common Stock",   indent="      ")
    ps_a,   ps_b,   _ps_e   = add_group("      Preferred Stock",            "Preferred Stock", indent="      ")
    apic_a, apic_b, _apic_e = add_group("      Additional Paid-In Capital", "APIC",           indent="      ")

    retained_a = total_assets_a - total_liab_a - cs_a - ps_a - apic_a - net_income
    retained_b = (total_assets_b - total_liab_b - cs_b - ps_b - apic_b - net_income_b) if has_b else None
    row("      Retained Earnings", retained_a, retained_b, None, "equity_item")
    row("      Net Income (Period)", net_income, net_income_b if has_b else None, None, "equity_item")

    total_equity_a = cs_a + ps_a + apic_a + retained_a + net_income
    total_equity_b = (cs_b + ps_b + apic_b + (retained_b or 0.0) + net_income_b) if has_b else None
    row("   Total Equity", total_equity_a, total_equity_b, None, "total")

    total_l_and_e_a = total_liab_a + total_equity_a
    total_l_and_e_b = (total_liab_b + (total_equity_b or 0.0)) if has_b else None
    # Equity has no IC eliminations; only liabilities carry IC payable elim
    total_l_and_e_e = total_liab_e
    row("TOTAL LIABILITIES AND EQUITY", total_l_and_e_a, total_l_and_e_b, total_l_and_e_e or None, "grand_total")

    row(None, None, None, None, "blank")
    check = total_assets_a - total_l_and_e_a
    R.append(("Check (Assets − Liabilities & Equity)", check, None, None, "check"))

    return R


# ── Excel writer ──────────────────────────────────────────────────────────────

def _build_bsi_excel(rows: list, as_of: str, company_name: str) -> bytes:
    """Render BS (Individual) Excel from a list of (label, value, rtype) tuples."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BS (Individual)"

    # ── Styles ───────────────────────────────────────────────────────────────
    def F(bold=False, size=9, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    DARK_NAVY  = PatternFill("solid", fgColor="0F172A")
    DARK_BLUE  = PatternFill("solid", fgColor="1E3A5F")
    MED_BLUE   = PatternFill("solid", fgColor="EBF2FB")
    LIGHT_GRAY = PatternFill("solid", fgColor="F1F5F9")
    GREEN_FILL = PatternFill("solid", fgColor="F0FDF4")
    RED_FILL   = PatternFill("solid", fgColor="FEF2F2")
    SUBSEC_BG  = PatternFill("solid", fgColor="F8FAFC")

    AL = Alignment(horizontal="left",   vertical="center")
    AR = Alignment(horizontal="right",  vertical="center")
    AC = Alignment(horizontal="center", vertical="center")

    thin  = Side(border_style="thin",   color="CBD5E1")
    thick = Side(border_style="medium", color="64748B")
    TOP1  = Border(top=thin)
    TOP2  = Border(top=thick)

    NUM = "#,##0.00"

    # ── Title block (mirrors Excel rows 1-3) ─────────────────────────────────
    ws.append([company_name, None, None, None, None])
    ws["A1"].font = F(bold=True, size=14, color="0F172A"); ws.row_dimensions[1].height = 22

    ws.append(["Balance Sheet", None, None, None, None])
    ws["A2"].font = F(bold=True, size=11, color="1E3A5F")

    ws.append([f"As of {as_of}", None, None, None, None])
    ws["A3"].font = F(size=9, color="64748B")

    ws.append([None] * 5)  # spacer

    # ── Column header row (mirrors Excel row 5) ───────────────────────────────
    headers = ["Particulars", company_name,
               "Concertiv Insurance Brokers, Inc.", "Eliminations", "Consolidated BS"]
    ws.append(headers)
    for ci in range(1, 6):
        c = ws.cell(5, ci)
        c.font      = F(bold=True, size=9, color="FFFFFF")
        c.fill      = DARK_BLUE
        c.alignment = AL if ci == 1 else AC
    ws.row_dimensions[5].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_tuple in rows:
        # Support 5-tuple (label, co_a, co_b, elim, rtype), 4-tuple, or legacy 3-tuple
        if len(row_tuple) == 5:
            label, value, co_b_val, elim_val, rtype = row_tuple
        elif len(row_tuple) == 4:
            label, value, co_b_val, rtype = row_tuple
            elim_val = None
        else:
            label, value, rtype = row_tuple
            co_b_val = None
            elim_val = None

        if rtype == "blank":
            ws.append([None] * 5)
            continue

        has_any = value is not None or co_b_val is not None
        cons = (value or 0.0) + (co_b_val or 0.0) + (elim_val or 0.0)
        ws.append([label, value, co_b_val if co_b_val is not None else None,
                   elim_val if elim_val is not None else None,
                   cons if has_any else None])
        r  = ws.max_row
        ca = ws.cell(r, 1)   # Particulars
        cb = ws.cell(r, 2)   # company_name
        cc = ws.cell(r, 3)   # broker
        cd = ws.cell(r, 4)   # Eliminations
        ce = ws.cell(r, 5)   # Consolidated

        ca.alignment = AL
        for cell in (cb, cc, cd, ce):
            if cell.value is not None:
                cell.number_format = NUM
                cell.alignment     = AR

        def fill_all(fill):
            for ci in range(1, 6): ws.cell(r, ci).fill = fill

        if rtype == "section":
            fill_all(DARK_NAVY)
            ca.font = F(bold=True, size=9, color="FFFFFF")
            for ci in range(2, 6): ws.cell(r, ci).font = F(bold=True, size=9, color="FFFFFF")
            ws.row_dimensions[r].height = 15

        elif rtype == "subsection":
            fill_all(SUBSEC_BG)
            ca.font = F(bold=True, size=9, color="1E293B")

        elif rtype == "group_header":
            fill_all(MED_BLUE)
            ca.font = F(bold=True, size=9, color="1E40AF")
            for ci in range(2, 6): ws.cell(r, ci).font = F(bold=True, size=9, color="1E40AF")
            ws.row_dimensions[r].height = 14

        elif rtype in ("account", "equity_item"):
            ca.font = F(size=9, color="334155")
            val_color_a = "DC2626" if value is not None and value < 0 else "334155"
            val_color_b = "DC2626" if co_b_val is not None and co_b_val < 0 else "334155"
            val_color_d = "DC2626" if elim_val is not None and elim_val < 0 else "334155"
            val_color_e = "DC2626" if cons < 0 else "334155"
            cb.font = F(size=9, color=val_color_a)
            cc.font = F(size=9, color=val_color_b)
            cd.font = F(size=9, color=val_color_d)
            ce.font = F(size=9, color=val_color_e)

        elif rtype == "group_total":
            fill_all(MED_BLUE)
            for ci in range(1, 6):
                c = ws.cell(r, ci)
                c.font   = F(bold=True, size=9, color="1E40AF")
                c.border = TOP1

        elif rtype == "total":
            fill_all(LIGHT_GRAY)
            for ci in range(1, 6):
                c = ws.cell(r, ci)
                c.font   = F(bold=True, size=9)
                c.border = TOP1

        elif rtype == "grand_total":
            fill_all(DARK_BLUE)
            for ci in range(1, 6):
                c = ws.cell(r, ci)
                c.font   = F(bold=True, size=10, color="FFFFFF")
                c.border = TOP2
            ws.row_dimensions[r].height = 16

        elif rtype == "check":
            is_zero = value is not None and abs(value) < 0.01
            fill_all(GREEN_FILL if is_zero else RED_FILL)
            txt = "166534" if is_zero else "991B1B"
            ca.font = F(bold=True, size=9, color=txt)
            cb.font = F(bold=True, size=9, color=txt)
            ce.font = F(bold=True, size=9, color=txt)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "B6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_bs_individual(df: pd.DataFrame, company_name: str) -> bytes:
    """Generate BS Individual Excel for the latest month from a raw DataFrame."""
    acct_a, net_a, acct_b, net_b, as_of = _compute_split(df)
    rows = _build_rows(acct_a, net_a, acct_b, net_b)
    return _build_bsi_excel(rows, as_of, company_name)


def run_bs_individual_from_preview(preview: dict, month: str, company_name: str) -> bytes:
    """Generate BS Individual Excel for a specific month from the cached preview JSONB."""
    rows_by_month = preview.get("rows_by_month", {})
    months        = preview.get("months", [])

    selected = month if month in rows_by_month else (months[-1] if months else "")
    raw_rows  = rows_by_month.get(selected, [])

    # Map {label, co_a, co_b, elim, type} → (label, co_a, co_b, elim, rtype)
    rows = [(r.get("label"), r.get("co_a"), r.get("co_b") or None, r.get("elim") or None, r.get("type")) for r in raw_rows]
    return _build_bsi_excel(rows, selected, company_name)


# ── Preview payload ───────────────────────────────────────────────────────────

def get_bs_individual_preview(df: pd.DataFrame) -> dict:
    """
    Return JSON-serialisable preview for the frontend.

    Shape:
      {
        months:        ["Jan-24", ...],
        as_of:         "Feb-26",        (latest available month)
        rows_by_month: {
          "Feb-26": [
            {label, co_a, co_b, elim, cons, type},
            ...
          ],
          ...
        }
      }

    co_b (Insurance Brokers) is populated when df contains broker rows (tagged
    with _source='broker').  elim (Eliminations) is always 0.
    cons (Consolidated) = co_a + co_b.
    """
    bs_df = df[df["_Financials"] == "Balance Sheet"].copy()
    all_bs_months = _sort_months(bs_df["_Month"].dropna().unique().tolist())

    if not all_bs_months:
        return {"months": [], "as_of": "N/A", "rows_by_month": {}}

    rows_by_month: dict = {}
    for month in all_bs_months:
        acct_a, net_a, acct_b, net_b, _ = _compute_split(df, as_of_month=month)
        raw_rows = _build_rows(acct_a, net_a, acct_b, net_b)
        rows_by_month[month] = [
            {
                "label": lbl,
                "co_a":  round(val_a, 2) if val_a is not None else None,
                "co_b":  round(val_b, 2) if val_b is not None else None,
                "elim":  round(elim, 2) if elim is not None else None,
                "cons":  round((val_a or 0.0) + (val_b or 0.0) + (elim or 0.0), 2)
                         if (val_a is not None or val_b is not None) else None,
                "type":  rt,
            }
            for lbl, val_a, val_b, elim, rt in raw_rows
        ]

    return {
        "months":        all_bs_months,
        "as_of":         all_bs_months[-1],
        "rows_by_month": rows_by_month,
    }
