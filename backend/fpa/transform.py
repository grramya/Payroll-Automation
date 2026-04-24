import io
import pandas as pd
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .mapping_data import ACCOUNT_MAP, DEPT_MAP
from .base_bs import run_base_bs, get_bs_preview
from .bs_individual import run_bs_individual, get_bs_individual_preview
from .pl_individual import run_pl_individual, get_pl_individual_preview
from .comparative_pl import run_comparative_pl, get_comparative_pl_preview
from .comparative_pl_bd import run_comparative_pl_bd, get_comparative_pl_bd_preview

# ── Alias: Acme Insurance variant → Concertiv Insurance entry ────────────────
_ACME = "130080 Intercompany Receivable - due from Acme Insurance"
_CONC = "130080 Intercompany Receivable - due from Concertiv Insurance"
if _ACME not in ACCOUNT_MAP and _CONC in ACCOUNT_MAP:
    ACCOUNT_MAP[_ACME] = ACCOUNT_MAP[_CONC]
for dept_key in list(DEPT_MAP.keys()):
    if dept_key[0] == _CONC:
        new_key = (_ACME, dept_key[1])
        if new_key not in DEPT_MAP:
            DEPT_MAP[new_key] = DEPT_MAP[dept_key]

# Pre-flatten ACCOUNT_MAP: one dict per field avoids double .get() on every row
_FIN_MAP   = {k: v.get("Financial Statement")         for k, v in ACCOUNT_MAP.items()}
_MAIN_MAP  = {k: v.get("Main Grouping")               for k, v in ACCOUNT_MAP.items()}
_SEC_MAP   = {k: v.get("Secondary Grouping")          for k, v in ACCOUNT_MAP.items()}
_CLASS_MAP = {k: v.get("Classification (Line Item)")  for k, v in ACCOUNT_MAP.items()}

# Pre-build reverse index: account → [(dept, vals)] so dept fallback is O(matches)
# instead of O(entire DEPT_MAP) per row
_DEPT_BY_ACCT: dict = defaultdict(list)
for (_d_acct, _d_dept), _d_vals in DEPT_MAP.items():
    _DEPT_BY_ACCT[_d_acct].append((_d_dept, _d_vals))

# ── Class (QuickBooks hierarchy) → Department (Class) ────────────────────────
CLASS_TO_DEPT = {
    "COGS":                                         "Cost of Revenue",
    "COGS:Procurement":                             "Cost of Revenue",
    "COGS:Procurement:Insurance":                   "Cost of Revenue",
    "COGS:Procurement:Tech & MD:Market Data":       "Cost of Revenue",
    "COGS:Procurement:Tech & MD:Readiness and Analysis": "Cost of Revenue",
    "COGS:Procurement:Tech & MD:Technology":        "Cost of Revenue",
    "COGS:Procurement:Travel":                      "Cost of Revenue",
    "COGS:Shared Services":                         "Cost of Revenue",
    "COGS:Shared Services:Day to Day Support":      "Cost of Revenue",
    "Client Service:Client Expert":                 "Customer Success",
    "Client Service:Client Management":             "Customer Success",
    "Client Service:Client Success":                "Customer Success",
    "Client Service:Procurement":                   "Customer Success",
    "Client Success (2023)":                        "Customer Success",
    "G&A":                                          "G&A",
    "G&A:Admin":                                    "G&A",
    "G&A:Management":                               "G&A",
    "G&A:Operations":                               "G&A",
    "R&D":                                          "R&D",
    "R&D:Engineering":                              "R&D",
    "R&D:Product":                                  "Product",
    "S&M:Marketing":                                "Marketing",
    "S&M:Sales":                                    "Sales",
    "S&M:Sales:Business Development":               "Sales",
    "S&M:Sales:Sales":                              "Sales",
}


def get_file_meta(input_bytes: bytes) -> dict:
    """Extract company name from the QuickBooks export metadata rows."""
    raw = pd.read_excel(io.BytesIO(input_bytes), header=None)
    company_name = ""
    for idx, row in raw.iterrows():
        if any(str(v).strip() == "Account" for v in row.values):
            break
        if idx == 0:
            for v in row.values:
                s = str(v).strip()
                if s and s.lower() != "nan":
                    company_name = s
                    break
    return {"company_name": company_name}


def _sv(v):
    """Return None for NaN/NaT, else the value as-is."""
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    return v


def _resolve_name(row) -> str | None:
    """Return the best available name for a transaction row.

    Fall back chain: Name → Vendor → Customer. All three columns are
    optional — absent columns are treated as None.
    """
    for col in ("Name", "Vendor", "Customer"):
        val = _sv(row[col]) if col in row.index else None
        if val is not None:
            return val
    return None


def run_transform(input_bytes: bytes, company_name: str = "Acme Corp, Inc.") -> tuple[bytes, dict, list]:
    # QuickBooks exports often include metadata rows (company name, report
    # title, date range) before the real column headers.  Read without a
    # header first, find the row that contains "Account", then re-read using
    # that row as the header so the DataFrame columns are correct.
    raw = pd.read_excel(io.BytesIO(input_bytes), header=None)
    header_row = None
    for idx, row in raw.iterrows():
        if any(str(v).strip() == "Account" for v in row.values):
            header_row = idx
            break
    if header_row is None:
        raise ValueError(
            "Could not find an 'Account' column in the uploaded file. "
            "Please upload a QuickBooks Transaction Detail report exported as .xlsx."
        )
    # Slice from the already-read raw frame instead of re-reading the file
    df = raw.iloc[header_row + 1:].copy()
    df.columns = raw.iloc[header_row].values
    df.columns = df.columns.astype(str).str.strip()

    # Drop any fully-empty rows that sometimes follow the header in QB exports
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Helper: return column series if it exists, else a None-filled series
    def col(name):
        return df[name] if name in df.columns else pd.Series([None] * len(df), index=df.index)

    # ── Normalize account and class ───────────────────────────────────────────
    df["Account"] = df["Account"].astype(str).str.strip()

    df["_ClassStr"] = col("Class").apply(
        lambda v: None if pd.isna(v) else str(v).strip()
    )

    # Translate QB Class → Department (Class)
    df["_DeptClass"] = df["_ClassStr"].map(lambda c: CLASS_TO_DEPT.get(c) if c else None)

    # ── Dates ─────────────────────────────────────────────────────────────────
    df["_date"]    = pd.to_datetime(col("Date"), errors="coerce")
    df["_Month"]   = df["_date"].dt.strftime("%b-%y")
    df["_Quarter"] = df["_date"].apply(
        lambda dt: f"Q{(dt.month - 1) // 3 + 1}-{dt.year}" if pd.notna(dt) else None
    )

    # ── Account-only lookups (Financials / Grouping / Classification 1) ───────
    df["_Financials"]        = df["Account"].map(_FIN_MAP)
    df["_MainGrouping"]      = df["Account"].map(_MAIN_MAP)
    df["_SecondaryGrouping"] = df["Account"].map(_SEC_MAP)
    df["_Classification"]    = df["Account"].map(_CLASS_MAP)

    # ── Composite (account, dept_class) lookups ───────────────────────────────
    # Single pass over rows; reverse index makes the fallback O(acct entries) not O(DEPT_MAP)
    _DEPT_FIELDS = ["Classification 2", "Classification 3", "Department (Class)", "Department Group (BD)"]

    def dept_lookup_all(row):
        acct = row["Account"]
        dept = row["_DeptClass"]
        results = [None, None, None, None]
        for i, field in enumerate(_DEPT_FIELDS):
            val = DEPT_MAP.get((acct, dept), {}).get(field)
            if val is None:
                val = DEPT_MAP.get((acct, None), {}).get(field)
            if val is None:
                for _, vals in _DEPT_BY_ACCT.get(acct, []):
                    if vals.get(field) is not None:
                        val = vals[field]
                        break
            results[i] = val
        return results

    dept_cols = df.apply(dept_lookup_all, axis=1, result_type="expand")
    df["_Classification2"] = dept_cols[0]
    df["_Classification3"] = dept_cols[1]
    df["_DeptClassOut"]    = dept_cols[2]
    df["_DeptGroupBD"]     = dept_cols[3]

    def clean_id(v):
        if pd.isna(v): return None
        try: return int(v)
        except: return v

    df["_AccountID"] = col("Account ID").apply(clean_id)

    # ── Summary ───────────────────────────────────────────────────────────────
    total   = len(df)
    matched = int(df["_Financials"].notna().sum())

    unmatched_df    = df[df["_Financials"].isna()]
    unmatched_accts = []
    if not unmatched_df.empty and "Account" in unmatched_df.columns:
        for acct, grp in unmatched_df.groupby("Account", dropna=False):
            label = "(Blank / missing Account)" if str(acct) == "nan" else str(acct)
            if "Amount" in grp.columns:
                raw = pd.to_numeric(grp["Amount"], errors="coerce").sum()
                amount = 0.0 if pd.isna(raw) else round(float(raw), 2)
            else:
                amount = 0.0
            unmatched_accts.append({"account": label, "rows": int(len(grp)), "amount": amount})

    fin_dist = {
        str(k): int(v)
        for k, v in df["_Financials"].value_counts(dropna=False).items()
    }
    date_min, date_max = df["_date"].min(), df["_date"].max()
    date_range = (
        f"{date_min.strftime('%B %Y')} - {date_max.strftime('%B %Y')}"
        if pd.notna(date_min) else "Unknown"
    )

    # ── Preview: expense P&L rows first (have dept data), then revenue, then BS ──
    expense_rows = df[df["_DeptClassOut"].notna()].head(50)
    other_rows   = df[df["_DeptClassOut"].isna()].head(50)
    preview_df   = pd.concat([expense_rows, other_rows]).reset_index(drop=True)

    def _rc(row, name):
        """Safely read an optional column from a row; return None if absent."""
        return _sv(row[name]) if name in row.index else None

    preview = []
    for _, row in preview_df.iterrows():
        preview.append({
            "Company":            company_name,
            "Date":               _rc(row, "Date"),
            "Transaction Type":   _rc(row, "Transaction Type"),
            "Num":                _rc(row, "Num"),
            "Name":               _resolve_name(row),
            "Class":              _rc(row, "Class"),
            "Memo/Description":   _rc(row, "Memo/Description"),
            "Split":              _rc(row, "Split"),
            "Amount":             _rc(row, "Amount"),
            "Balance":            _rc(row, "Balance"),
            "Vendor":             _rc(row, "Vendor"),
            "Customer":           _rc(row, "Customer"),
            "Account":            _sv(row["Account"]),
            "Account ID":         _sv(row["_AccountID"]),
            "Financials":         _sv(row["_Financials"]),
            "Main Grouping":      _sv(row["_MainGrouping"]),
            "Secondary Grouping": _sv(row["_SecondaryGrouping"]),
            "Classification":     _sv(row["_Classification"]),
            "Month":              _sv(row["_Month"]),
            "Classification 2":   _sv(row["_Classification2"]),
            "Classification 3":   _sv(row["_Classification3"]),
            "Class (Dept)":       _sv(row["_DeptClassOut"]),
            "Class Group (BD)":   _sv(row["_DeptGroupBD"]),
            "One time Expenses":  None,
            "Quarter":            _sv(row["_Quarter"]),
        })

    # ── Build Excel (exact staging file format) ───────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    YELLOW  = PatternFill("solid", fgColor="FFFF00")
    NO_FILL = PatternFill(fill_type=None)
    F14B    = Font(name="Calibri", bold=True,  size=14)
    F10B    = Font(name="Calibri", bold=True,  size=10)
    F11     = Font(name="Calibri", bold=False, size=11)
    F9B     = Font(name="Calibri", bold=True,  size=9)
    F8      = Font(name="Calibri", bold=False, size=8)
    CTR     = Alignment(horizontal="center")
    LEFT    = Alignment(horizontal="left")
    RIGHT   = Alignment(horizontal="right")
    NUM_FMT = '#,##0.00\\ _€'

    ws.append([company_name]                                          + [None]*24); ws["A1"].font = F14B; ws.row_dimensions[1].height = 18
    ws.append(["Transaction Detail by  Account"]                      + [None]*24); ws["A2"].font = F14B; ws.row_dimensions[2].height = 18
    ws.append([date_range]                                            + [None]*24); ws["A3"].font = F10B
    ws.append(["Mention the company name from where data is flowing"] + [None]*24); ws["A4"].font = F11

    HEADERS = [
        "Company", "Date", "Transaction Type", "Num", "Name", "Class",
        "Memo/Description", "Split", "Amount", "Balance", "Vendor", "Customer",
        "Account", "Account ID",
        "Financials", "Main Grouping", "Secondary Grouping", "Classification",
        "Month", "Classification 2", "Classification 3", "Class",
        "Class Group (BD)", "One time Expenses", "Quarter",
    ]
    ws.append(HEADERS)
    for ci, _ in enumerate(HEADERS, 1):
        cell = ws.cell(5, ci)
        if ci == 1:
            cell.font = F11; cell.fill = NO_FILL
        elif ci <= 14:
            cell.font = F9B; cell.fill = NO_FILL; cell.alignment = CTR
        else:
            cell.font = F9B; cell.fill = YELLOW;  cell.alignment = CTR

    for _, row in df.iterrows():
        ws.append([
            company_name,
            _rc(row, "Date"),
            _rc(row, "Transaction Type"),
            _rc(row, "Num"),
            _resolve_name(row),
            _rc(row, "Class"),
            _rc(row, "Memo/Description"),
            _rc(row, "Split"),
            _rc(row, "Amount"),
            _rc(row, "Balance"),
            _rc(row, "Vendor"),
            _rc(row, "Customer"),
            _sv(row["Account"]),
            _sv(row["_AccountID"]),
            _sv(row["_Financials"]),
            _sv(row["_MainGrouping"]),
            _sv(row["_SecondaryGrouping"]),
            _sv(row["_Classification"]),
            _sv(row["_Month"]),
            _sv(row["_Classification2"]),
            _sv(row["_Classification3"]),
            _sv(row["_DeptClassOut"]),
            _sv(row["_DeptGroupBD"]),
            None,
            _sv(row["_Quarter"]),
        ])

    for ws_row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for cell in ws_row:
            ci = cell.column
            if ci == 1:       cell.font = F11
            elif ci <= 14:    cell.font = F8;  cell.alignment = LEFT
            else:             cell.font = F11
        for ci in (9, 10):
            c = ws_row[ci - 1]
            if c.value is not None:
                c.number_format = NUM_FMT; c.alignment = RIGHT

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)

    summary = {
        "total_rows":              total,
        "matched_rows":            matched,
        "unmatched_rows":          total - matched,
        "unmatched_accounts":      unmatched_accts,
        "financials_distribution": fin_dist,
        "date_range":              date_range,
    }

    # ── Generate all 5 reports in parallel (none depend on each other) ───────
    with ThreadPoolExecutor(max_workers=10) as ex:
        f_bs        = ex.submit(run_base_bs,                   df, company_name)
        f_bs_p      = ex.submit(get_bs_preview,                df)
        f_bsi       = ex.submit(run_bs_individual,             df, company_name)
        f_bsi_p     = ex.submit(get_bs_individual_preview,     df)
        f_pl        = ex.submit(run_pl_individual,             df, company_name)
        f_pl_p      = ex.submit(get_pl_individual_preview,     df, company_name)
        f_comp      = ex.submit(run_comparative_pl,            df, company_name)
        f_comp_p    = ex.submit(get_comparative_pl_preview,    df, company_name)
        f_comp_bd   = ex.submit(run_comparative_pl_bd,         df, company_name)
        f_comp_bd_p = ex.submit(get_comparative_pl_bd_preview, df, company_name)

    bs_bytes,         bs_preview         = f_bs.result(),      f_bs_p.result()
    bsi_bytes,        bsi_preview        = f_bsi.result(),     f_bsi_p.result()
    pl_bytes,         pl_preview         = f_pl.result(),      f_pl_p.result()
    comp_pl_bytes,    comp_pl_preview    = f_comp.result(),    f_comp_p.result()
    comp_pl_bd_bytes, comp_pl_bd_preview = f_comp_bd.result(), f_comp_bd_p.result()

    return (
        buf.getvalue(), summary, preview,
        bs_bytes, bs_preview,
        bsi_bytes, bsi_preview,
        pl_bytes, pl_preview,
        comp_pl_bytes, comp_pl_preview,
        comp_pl_bd_bytes, comp_pl_bd_preview,
    )
