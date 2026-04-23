"""
Comparative P&L (BD) generator.

Mirrors the "Comparative P&L (BD)" sheet in the Excel reference.

Column groups (identical to Excel layout):
  A        Particulars
  B C D    3 monthly columns  (3 months ending at the latest "as of" month)
  E        blank separator
  F G H I  4 quarterly columns — Q1–Q4 of (quarter_year − 1)
  J        blank separator
  K L M N  4 quarterly columns — Q1–Q4 of quarter_year
  O        blank separator
  P Q      2 yearly columns   — (year − 1) and year

Key difference vs Comparative P&L (Class):
  • Department grouping uses _DeptGroupBD (col W in Transaction Detail / Class List col D),
    NOT _DeptClassOut (col V / Class List col C).
  • BD collapses the granular classes into 3 buckets:
      Sales & Marketing  = Sales + Marketing + Customer Success
      Research & Dev.    = R&D + Product
      General & Admin.   = G&A
  • Cost of Goods Sold is a single aggregated line (not split into sub-items).
  • D&A is a BELOW-EBITDA deduction, not part of Operating Expenses.
    Therefore: EBITDA = Gross Margin − (S&M + R&D + G&A)  [exact formula, no D&A inside]
  • Net Income = EBITDA + Other_Income + Other_Expenses − D&A − Tax

Aggregation formula equivalents (mirror the Excel SUMIFS chains):

  Revenue   → SUMIFS(Amount, _Classification2 = label,           period_filter)
  COGS      → SUMIFS(Amount, _DeptGroupBD     = "Cost of Goods Sold", period_filter)
  S&M       → SUMIFS(Amount, _DeptGroupBD     = "Sales & Marketing",  period_filter)
  R&D       → SUMIFS(Amount, _DeptGroupBD     = "Research & Development", period_filter)
  G&A       → SUMIFS(Amount, _DeptGroupBD     = "General & Administrative", period_filter)
  D&A       → SUMIFS(Amount, _Classification2 = "Depreciation and Amortization", period_filter)
  Other Inc → SUMIFS(Amount, _Classification2 = "Other Income",    period_filter)
  Other Exp → SUMIFS(Amount, _Classification2 = "Other Expenses",  period_filter)
  Tax       → SUMIFS(Amount, _Classification2 = "Tax Expense",     period_filter)

Quarter labels use the same formula as the Excel Period Mapping sheet:
  "Q" & INT((MONTH-1)/3)+1 & "-" & YEAR  →  f"Q{(month-1)//3+1}-{year}"
  (identical to the _Quarter column produced by transform.py)
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .pl_individual import (
    REVENUE_LINES,
    _sort_months,
)
from .comparative_pl import _sort_quarters, _prev_months


# ── BD-specific constants ──────────────────────────────────────────────────────

# BD expense buckets (single aggregated line each, filtered by _DeptGroupBD)
BD_EXPENSE_DEPTS = [
    ("Sales & Marketing",        "Sales & Marketing"),
    ("Research & Development",   "Research & Development"),
    ("General & Administrative", "General & Administrative"),
]


# ── Row structure definition ───────────────────────────────────────────────────

def _build_row_structure() -> list[dict]:
    """
    Static row structure mirroring the Excel Comparative P&L (BD) layout.
    Each entry has:
      label  – display text
      type   – section | subsection | line | total | grand_total | metric | blank
      key    – data dict key (None for structural/computed-display-only rows)
    """
    R: list[dict] = []
    def row(label, rtype, key=None):
        R.append({"label": label, "type": rtype, "key": key})

    # ── Income ────────────────────────────────────────────────────────────────
    row("Income",  "section")
    row("Revenue", "subsection")
    for label, _, _ in REVENUE_LINES:
        row(f"   {label}", "line", label)
    row("Total Revenue", "total", "Total Revenue")
    row(None, "blank")

    # ── Cost of Goods Sold ────────────────────────────────────────────────────
    row("Cost of Goods Sold", "total", "Cost of Goods Sold")
    row(None, "blank")

    # ── Gross Margin ──────────────────────────────────────────────────────────
    row("Gross Margin",     "grand_total", "Gross Margin")
    row("Gross Margin (%)", "metric",      "Gross Margin (%)")
    row(None, "blank")

    # ── Expenses (S&M, R&D, G&A — D&A excluded here per BD convention) ───────
    row("Expenses", "section")
    for label, _ in BD_EXPENSE_DEPTS:
        row(f"   {label}", "line", label)
    row("Total Expenses", "total", "Total Expenses")
    row(None, "blank")

    # ── EBITDA ────────────────────────────────────────────────────────────────
    row("EBITDA",    "grand_total", "EBITDA")
    row("EBITDA (%)", "metric",     "EBITDA (%)")
    row(None, "blank")

    # ── Below-EBITDA items (deducted from EBITDA to arrive at Net Income) ────
    row("Other Income (Expenses)", "section")
    row("   Other Income",                    "line",  "Other Income")
    row("   Other Expenses",                  "line",  "Other Expenses")
    row("   Depreciation and Amortization",   "line",  "Depreciation and Amortization")
    row("   Tax Expense",                     "line",  "Tax Expense")
    row("   Total Below EBITDA",              "total", "Total Below EBITDA")
    row(None, "blank")

    # ── Net Income ────────────────────────────────────────────────────────────
    row("Net Income",     "grand_total", "Net Income")
    row("Net Income (%)", "metric",      "Net Income (%)")

    return R


ROW_STRUCTURE = _build_row_structure()


# ── Core aggregation ───────────────────────────────────────────────────────────

def _compute_period(m: pd.DataFrame) -> dict:
    """
    Compute all BD P&L values for a single period slice (month / quarter / year).
    Returns a flat dict of {data_key: float | None}.

    Formula chain (mirrors Excel exactly):
      Gross Margin  = Total Revenue − Cost of Goods Sold
      EBITDA        = Gross Margin  − Total Expenses            [D&A NOT in Expenses]
      Below EBITDA  = Other Income + Other Expenses + D&A + Tax [all with natural signs]
      Net Income    = EBITDA − Total Below EBITDA
    """
    d: dict = {}

    # ── Revenue ───────────────────────────────────────────────────────────────
    rev_total = 0.0
    client_recurring = 0.0
    for label, _, cls2 in REVENUE_LINES:
        val = float(m.loc[m["_Classification2"] == cls2, "Amount"].sum())
        d[label] = val
        rev_total += val
        if label == "Client Recurring":
            client_recurring = val
    d["Total Revenue"] = rev_total

    # ── Cost of Goods Sold (single BD bucket via _DeptGroupBD) ───────────────
    cogs = float(m.loc[m["_DeptGroupBD"] == "Cost of Goods Sold", "Amount"].sum())
    d["Cost of Goods Sold"] = cogs

    # ── Gross Margin = Revenue − COGS ─────────────────────────────────────────
    gm = rev_total - cogs
    d["Gross Margin"] = gm
    d["Gross Margin (%)"] = (
        (gm / rev_total * 100) if rev_total and abs(rev_total) > 0.001 else None
    )

    # ── Expenses — 3 BD buckets, D&A excluded (it goes below EBITDA) ─────────
    exp_total = 0.0
    for label, bd_key in BD_EXPENSE_DEPTS:
        val = float(m.loc[m["_DeptGroupBD"] == bd_key, "Amount"].sum())
        d[label] = val
        exp_total += val
    d["Total Expenses"] = exp_total

    # ── EBITDA = Gross Margin − Total Expenses ────────────────────────────────
    ebitda = gm - exp_total
    d["EBITDA"] = ebitda
    d["EBITDA (%)"] = (
        (ebitda / rev_total * 100) if rev_total and abs(rev_total) > 0.001 else None
    )

    # ── Below-EBITDA items ────────────────────────────────────────────────────
    oi  = float(m.loc[m["_Classification2"] == "Other Income",    "Amount"].sum())
    oe  = float(m.loc[m["_Classification2"] == "Other Expenses",  "Amount"].sum())
    da  = float(m.loc[m["_Classification2"] == "Depreciation and Amortization", "Amount"].sum())
    tax = float(m.loc[m["_Classification2"] == "Tax Expense",     "Amount"].sum())

    d["Other Income"]                  = oi
    d["Other Expenses"]                = oe
    d["Depreciation and Amortization"] = da
    d["Tax Expense"]                   = tax

    # Total Below EBITDA = Other Income + Other Expenses + D&A + Tax
    # (amounts carry natural signs: Other Income ≥ 0, Other Expenses ≤ 0, D&A ≥ 0, Tax ≥ 0)
    total_below = oi + oe + da + tax
    d["Total Below EBITDA"] = total_below

    # ── Net Income = EBITDA − Total Below EBITDA ──────────────────────────────
    # Mirrors Excel: =EBITDA_cell − SUM(below_items)
    ni = ebitda - total_below
    d["Net Income"] = ni
    d["Net Income (%)"] = (
        (ni / rev_total * 100) if rev_total and abs(rev_total) > 0.001 else None
    )

    return d


def _aggregate(df: pd.DataFrame) -> tuple[dict, list, list, list]:
    """
    Aggregate BD P&L data across all months, quarters, and years present in df.

    Returns:
      data     – {period_key: {data_key: float | None}}
      months   – chronologically sorted list of 'Mmm-yy' strings
      quarters – chronologically sorted list of 'Q{1-4}-{YYYY}' strings
      years    – sorted list of int years
    """
    pl = df[df["_Financials"] == "Profit and Loss A/c"].copy()
    pl = pl[pl["_Month"].notna() & pl["Amount"].notna()]

    if pl.empty:
        return {}, [], [], []

    all_months   = _sort_months(
        [m for m in df["_Month"].dropna().unique().tolist() if m]
    )
    all_quarters = _sort_quarters(
        [q for q in df["_Quarter"].dropna().unique().tolist() if q]
    )
    all_years = sorted(
        [int(y) for y in df["_date"].dt.year.dropna().unique()]
    )

    data: dict[str, dict] = {}

    for m_label in all_months:
        data[m_label] = _compute_period(pl[pl["_Month"] == m_label])

    for q_label in all_quarters:
        data[q_label] = _compute_period(pl[pl["_Quarter"] == q_label])

    for year in all_years:
        data[str(year)] = _compute_period(pl[pl["_date"].dt.year == year])

    return data, all_months, all_quarters, all_years


# ── Excel writer ───────────────────────────────────────────────────────────────

def run_comparative_pl_bd(df: pd.DataFrame, company_name: str) -> bytes:
    """
    Generate the Comparative P&L (BD) Excel workbook.

    Column layout (mirrors Excel reference):
      A         Particulars
      B C D     3 monthly columns (latest 3 months ending at as_of)
      E         blank separator
      F G H I   Q1–Q4 of (quarter_year − 1)
      J         blank separator
      K L M N   Q1–Q4 of quarter_year
      O         blank separator
      P Q       (year − 1) and year
    """
    data, months, quarters, years = _aggregate(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparative P&L (BD)"

    if not months:
        ws["A1"] = "No P&L data found in the uploaded file."
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    # ── Determine column periods ───────────────────────────────────────────────
    as_of        = months[-1]
    month_cols   = _prev_months(months, as_of, 3)

    latest_q     = quarters[-1] if quarters else None
    q_year       = int(latest_q[3:]) if latest_q else (years[-1] if years else datetime.now().year)
    prior_q_year = q_year - 1

    q_prior = [f"Q{n}-{prior_q_year}" for n in range(1, 5)]
    q_curr  = [f"Q{n}-{q_year}"       for n in range(1, 5)]

    cur_year  = years[-1] if years else datetime.now().year
    prev_year = cur_year - 1

    period_cols = (
        month_cols +
        [None] +
        q_prior +
        [None] +
        q_curr +
        [None] +
        [str(prev_year), str(cur_year)]
    )

    # ── Styles ────────────────────────────────────────────────────────────────
    def F(bold=False, size=9, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    DARK_NAVY  = PatternFill("solid", fgColor="0F172A")
    DARK_BLUE  = PatternFill("solid", fgColor="1E3A5F")
    MED_BLUE   = PatternFill("solid", fgColor="EBF2FB")
    LIGHT_GRAY = PatternFill("solid", fgColor="F1F5F9")
    SUBSEC_BG  = PatternFill("solid", fgColor="F8FAFC")
    METRIC_BG  = PatternFill("solid", fgColor="F0F9FF")
    SEP_FILL   = PatternFill("solid", fgColor="F8FAFC")

    AL = Alignment(horizontal="left",   vertical="center")
    AR = Alignment(horizontal="right",  vertical="center")
    AC = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin",   color="CBD5E1")
    NUM  = "#,##0.00"
    PCTF = "0.00%"

    n_data_cols = len(period_cols)

    # ── Title block ───────────────────────────────────────────────────────────
    ws.append([company_name] + [None] * n_data_cols)
    ws["A1"].font = F(bold=True, size=14, color="0F172A"); ws.row_dimensions[1].height = 22

    ws.append(["Comparative P&L (BD)"] + [None] * n_data_cols)
    ws["A2"].font = F(bold=True, size=11, color="1E3A5F")

    ws.append([f"As of {as_of}"] + [None] * n_data_cols)
    ws["A3"].font = F(size=9, color="64748B", italic=True)

    ws.append([None] * (n_data_cols + 1))

    # ── Group label row (row 5) ───────────────────────────────────────────────
    group_row = [None] * (n_data_cols + 1)
    group_row[1]  = "Months"
    group_row[5]  = f"Quarters {prior_q_year}"
    group_row[10] = f"Quarters {q_year}"
    group_row[15] = "Year"
    ws.append(group_row)
    gr = ws.max_row
    ws.row_dimensions[gr].height = 14

    merge_map = {1: 3, 5: 8, 10: 13, 15: 16}
    for col_start, col_end in merge_map.items():
        ws.merge_cells(
            start_row=gr, start_column=col_start + 1,
            end_row=gr,   end_column=col_end + 1,
        )
        c = ws.cell(gr, col_start + 1)
        c.font      = F(bold=True, size=8, color="FFFFFF")
        c.fill      = DARK_BLUE
        c.alignment = AC

    # ── Period header row (row 6) ─────────────────────────────────────────────
    hdr = ["Particulars"] + [p if p is not None else "" for p in period_cols]
    ws.append(hdr)
    hr = ws.max_row
    ws.row_dimensions[hr].height = 16
    for ci, val in enumerate(hdr, 1):
        c = ws.cell(hr, ci)
        if ci == 1:
            c.font = F(bold=True, size=9, color="FFFFFF"); c.fill = DARK_BLUE; c.alignment = AL
        elif val == "":
            c.fill = SEP_FILL
        else:
            c.font = F(bold=True, size=9, color="FFFFFF"); c.fill = DARK_BLUE; c.alignment = AC

    # ── Data rows ─────────────────────────────────────────────────────────────
    for rdef in ROW_STRUCTURE:
        label = rdef["label"]
        rtype = rdef["type"]
        key   = rdef.get("key")

        if rtype == "blank":
            ws.append([None] * (n_data_cols + 1)); continue

        if key is None:
            vals = [None] * n_data_cols
        else:
            vals = []
            for p in period_cols:
                if p is None:
                    vals.append(None)
                else:
                    vals.append(data.get(p, {}).get(key))

        ws.append([label] + vals)
        r = ws.max_row

        ca = ws.cell(r, 1)
        ca.alignment = AL

        def fill_row(fill):
            for ci in range(1, n_data_cols + 2):
                ws.cell(r, ci).fill = fill

        for ci, (p, v) in enumerate(zip(period_cols, vals), 2):
            cell = ws.cell(r, ci)
            if p is None:
                cell.fill = SEP_FILL; continue
            if v is None: continue
            if rtype == "metric":
                cell.number_format = PCTF
                cell.value = v / 100 if v is not None else v
            else:
                cell.number_format = NUM
            cell.alignment = AR

        if rtype == "section":
            fill_row(DARK_NAVY)
            for ci in range(1, n_data_cols + 2):
                ws.cell(r, ci).font = F(bold=True, size=9, color="FFFFFF")
            ws.row_dimensions[r].height = 15

        elif rtype == "subsection":
            fill_row(SUBSEC_BG)
            ca.font = F(bold=True, size=9, color="1E293B")

        elif rtype == "line":
            ca.font = F(size=9, color="334155")
            for ci, p in enumerate(period_cols, 2):
                if p is None: continue
                cell = ws.cell(r, ci)
                if cell.value is not None:
                    neg = isinstance(cell.value, (int, float)) and cell.value < 0
                    cell.font = F(size=9, color="DC2626" if neg else "334155")

        elif rtype == "total":
            fill_row(LIGHT_GRAY)
            for ci in range(1, n_data_cols + 2):
                c = ws.cell(r, ci)
                c.font   = F(bold=True, size=9)
                c.border = Border(top=thin)

        elif rtype == "grand_total":
            fill_row(DARK_BLUE)
            for ci in range(1, n_data_cols + 2):
                c   = ws.cell(r, ci)
                val = c.value
                neg = isinstance(val, (int, float)) and val < 0
                c.font   = F(bold=True, size=10, color="FCA5A5" if neg else "FFFFFF")
                c.border = Border(top=Side(border_style="medium", color="64748B"))
            ws.row_dimensions[r].height = 16

        elif rtype == "metric":
            fill_row(METRIC_BG)
            ca.font = F(italic=True, size=8, color="0369A1")
            for ci in range(2, n_data_cols + 2):
                c = ws.cell(r, ci)
                if c.value is not None:
                    neg = isinstance(c.value, (int, float)) and c.value < 0
                    c.font = F(italic=True, size=8, color="DC2626" if neg else "0369A1")

    # ── Column widths & freeze ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    for col_idx, p in enumerate(period_cols, 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 2 if p is None else 14

    ws.freeze_panes = "B7"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Preview payload ────────────────────────────────────────────────────────────

def get_comparative_pl_bd_preview(df: pd.DataFrame, company_name: str = "") -> dict:
    """
    Return JSON-serialisable preview for the frontend.

    Shape:
      {
        company_name:        str,
        available_months:    ["Jan-24", ...],
        available_quarters:  ["Q1-2024", ...],
        available_years:     [2024, 2025, ...],
        data: {
          "Jan-24":  { data_key: value, ... },
          "Q1-2024": { data_key: value, ... },
          "2024":    { data_key: value, ... },
          ...
        },
        rows: [ { label, type, key }, ... ]
      }
    """
    data, months, quarters, years = _aggregate(df)

    if not months:
        return {
            "company_name":       company_name,
            "available_months":   [],
            "available_quarters": [],
            "available_years":    [],
            "data":               {},
            "rows":               ROW_STRUCTURE,
        }

    def _round_val(v):
        if v is None: return None
        return round(v, 2) if isinstance(v, float) else v

    serialisable_data: dict[str, dict] = {}
    for period_key, period_vals in data.items():
        serialisable_data[period_key] = {k: _round_val(v) for k, v in period_vals.items()}

    return {
        "company_name":       company_name,
        "available_months":   months,
        "available_quarters": quarters,
        "available_years":    years,
        "data":               serialisable_data,
        "rows":               ROW_STRUCTURE,
    }
