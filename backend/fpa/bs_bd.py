"""
Balance Sheet (BD) generator.

Column layout mirrors the BS (BD) 1.xlsx reference template:
  Q1–Q4 of selected year  |  Prior-year annual  |  Current-year annual

Mapping:
  Uses the "Allocation (BD)" column from the mapping table to group
  balance-sheet accounts into BD line items.  For ambiguous cases where the
  same Classification maps to multiple AllocBD values (Intangible assets →
  Customer relationships / Tradename / Developed Technology; Preferred Stock →
  Series B / B-1 / B-2) the account-number prefix is used to differentiate.

Balance logic (same as base_bs):
  Month-end snapshot = last Balance value per account in that month.
  Quarterly snapshot = last month available in the quarter (ffilled).
  Annual snapshot    = last month available in December (Q4) of that year.

Retained earnings is a plug so the sheet ties:
  = Total Assets − Current Liabilities − Non-current liabilities
    − Common stock − Preferred Series B − Preferred Series B-1
    − Preferred Series B-2 − Additional paid in capital
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── AllocBD lookup tables ─────────────────────────────────────────────────────

# Account-number-prefix → AllocBD (overrides classification-level for ambiguous cases)
_ACCT_ALLOC_BD: dict[str, str] = {
    "151000": "Customer relationships",
    "151010": "Customer relationships",
    "152000": "Developed Technology",
    "152010": "Developed Technology",
    "156000": "Tradename",
    "156010": "Tradename",
    "310070": "Preferred Series B",
    "310080": "Preferred Series B-1",
    "310090": "Preferred Series B-2",
}

# Classification (Line Item) → AllocBD (used when no account-level override)
_CLS_ALLOC_BD: dict[str, str] = {
    "Accounts Receivable":                                   "Accounts receivable",
    "Cash and Cash equivalents":                             "Cash",
    "Prepaid expenses and other current assets":             "Prepaid expenses and other current assets",
    "Other Current Assets":                                  "Prepaid expenses and other current assets",
    "Deposit or Advances":                                   "Prepaid expenses and other current assets",
    "Finance lease right-of-use assets":                     "ROU-Finance Lease",
    "Goodwill, net of accumulated amortization":             "Goodwill",
    "Operating lease right-of-use assets":                   "ROU-Operating Leases",
    "Tangible Assets, net of accumulated depreciation":      "Fixed assets",
    "Accounts Payable":                                      "Accounts payable",
    "Accrued expenses":                                      "Accrued expenses",
    "Due to Employee's":                                     "Accrued expenses",
    "Current portion of operating lease liabilities":        "Lease liabilities",
    "Deferred Revenue":                                      "Deferred revenue",
    "Statutory Dues":                                        "Non-current liabilities",
    "Finance lease liabilities, net of current portion":     "Non-current liabilities",
    "Operating lease liabilities, net of current portion":   "Non-current liabilities",
    "Other Long term Liabilities":                           "Non-current liabilities",
    "Additional paid-in capital":                            "Additional paid in capital",
    "Common Stock":                                          "Common stock",
    "Retained Earning":                                      "Retained earnings",
    "Preferred Stock":                                       "Preferred Series B",  # fallback
}


def _resolve_alloc_bd(account: str, classification: str) -> str | None:
    """Return the AllocBD label for a single row."""
    num = str(account).split()[0] if account else ""
    override = _ACCT_ALLOC_BD.get(num)
    if override:
        return override
    return _CLS_ALLOC_BD.get(classification)


# ── Row structure ─────────────────────────────────────────────────────────────
# (display_label, row_type, alloc_bd_key_or_children)
#
# row_type:
#   section    – dark header band (no numeric values)
#   subsection – lighter header band
#   data       – leaf line item; alloc_bd_key_or_children = AllocBD string
#   subtotal   – sum of listed data-key children
#   total      – sum of listed subtotal/group keys
#   retained   – plug calculation
#   check      – Total Assets − Total L&SE

BS_BD_ROWS: list[tuple] = [
    # ── Assets ────────────────────────────────────────────────────────────────
    ("Assets",                                                   "section",    None),
    ("Current Assets",                                           "subsection", None),
    ("Cash",                                                     "data",       "Cash"),
    ("Accounts receivable",                                      "data",       "Accounts receivable"),
    ("Prepaid expenses and other current assets",                "data",       "Prepaid expenses and other current assets"),
    ("Total Current Assets",                                     "subtotal",   ["Cash", "Accounts receivable", "Prepaid expenses and other current assets"]),
    (None,                                                       "blank",      None),
    ("Non-current Assets",                                       "subsection", None),
    ("Customer relationships",                                   "data",       "Customer relationships"),
    ("Goodwill",                                                 "data",       "Goodwill"),
    ("Tradename",                                                "data",       "Tradename"),
    ("Fixed assets",                                             "data",       "Fixed assets"),
    ("ROU-Operating Leases",                                     "data",       "ROU-Operating Leases"),
    # Developed Technology and ROU-Finance Lease flow into the total but are
    # not shown as separate lines in the template (kept as hidden contributors).
    ("Total Non-current Assets",                                 "subtotal",   [
        "Customer relationships", "Goodwill", "Tradename", "Fixed assets",
        "ROU-Operating Leases", "ROU-Finance Lease", "Developed Technology",
    ]),
    ("Total Assets",                                             "total",      ["Total Current Assets", "Total Non-current Assets"]),
    (None,                                                       "blank",      None),
    # ── Liabilities and Stockholders' Equity ──────────────────────────────────
    ("Liabilities and Stockholders' Equity",                     "section",    None),
    ("Current liabilities",                                      "subsection", None),
    ("Accounts payable",                                         "data",       "Accounts payable"),
    ("Accrued expenses",                                         "data",       "Accrued expenses"),
    ("Deferred revenue",                                         "data",       "Deferred revenue"),
    ("Lease liabilities",                                        "data",       "Lease liabilities"),
    ("Total Current Liabilities",                                "subtotal",   ["Accounts payable", "Accrued expenses", "Deferred revenue", "Lease liabilities"]),
    (None,                                                       "blank",      None),
    ("Non-current liabilities",                                  "data",       "Non-current liabilities"),
    (None,                                                       "blank",      None),
    ("Stockholders' equity",                                     "subsection", None),
    ("Common stock",                                             "data",       "Common stock"),
    ("Preferred Series B",                                       "data",       "Preferred Series B"),
    ("Preferred Series B-1",                                     "data",       "Preferred Series B-1"),
    ("Preferred Series B-2",                                     "data",       "Preferred Series B-2"),
    ("Additional paid in capital",                               "data",       "Additional paid in capital"),
    ("Retained earnings",                                        "retained",   None),
    ("Total stockholders' equity",                               "subtotal",   [
        "Common stock", "Preferred Series B", "Preferred Series B-1",
        "Preferred Series B-2", "Additional paid in capital", "Retained earnings",
    ]),
    ("Total liabilities and stockholders' equity",               "total",      [
        "Total Current Liabilities", "Non-current liabilities", "Total stockholders' equity",
    ]),
    (None,                                                       "blank",      None),
    ("Check",                                                    "check",      None),
]


# ── Value computation ─────────────────────────────────────────────────────────

_MONTH_ABBR = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def _month_to_ym(m: str) -> str:
    try:
        mon, yr = m.split("-")
        full_yr = f"20{yr}" if len(yr) == 2 else yr
        return f"{full_yr}-{_MONTH_ABBR[mon]}"
    except Exception:
        return "0000-00"


def _month_to_quarter(m: str) -> str:
    """'Jan-25' → 'Q1-2025'"""
    try:
        dt = datetime.strptime(m, "%b-%y")
        return f"Q{(dt.month - 1) // 3 + 1}-{dt.year}"
    except Exception:
        return ""


def _sort_months(months: list[str]) -> list[str]:
    return sorted(months, key=_month_to_ym)


def _sort_quarters(quarters: list[str]) -> list[str]:
    """Sort quarter strings like 'Q1-2025' chronologically."""
    def _q_key(q: str) -> str:
        try:
            return f"{q[3:]}-{q[1]}"  # "2025-1"
        except Exception:
            return q
    return sorted(quarters, key=_q_key)


def _build_monthly_balances(df: pd.DataFrame) -> tuple[dict[str, dict[str, float]], list[str]]:
    """
    Returns ({alloc_bd: {month: balance}}, sorted_months).

    Month-end balance per AllocBD = sum of last Balance values across all
    accounts that map to that AllocBD for that month (forward-filled for gaps).
    """
    bs = df[df["_Financials"] == "Balance Sheet"].copy()
    if bs.empty:
        return {}, []

    bs["_date_parsed"] = pd.to_datetime(bs["Date"], errors="coerce")
    bs = bs.sort_values("_date_parsed")

    # Resolve AllocBD for each row
    bs["_AllocBD"] = bs.apply(
        lambda r: _resolve_alloc_bd(
            str(r.get("Account", "")),
            str(r.get("_Classification", "")),
        ),
        axis=1,
    )

    txn_months = _sort_months(
        [m for m in bs["_Month"].dropna().unique() if m]
    )
    if not txn_months:
        return {}, []

    # Expand to a full calendar month range so every end-of-quarter month exists
    # even if there were no transactions in that specific month.
    first_ym = _month_to_ym(txn_months[0])
    last_ym  = _month_to_ym(txn_months[-1])
    start    = pd.Period(first_ym, freq="M")
    end      = pd.Period(last_ym,  freq="M")
    all_periods = pd.period_range(start, end, freq="M")
    all_months  = [p.strftime("%b-%y") for p in all_periods]

    # Month-end balance per (Account, AllocBD, Month)
    grp = (
        bs[bs["_Month"].notna() & bs["Balance"].notna()]
        .groupby(["Account", "_AllocBD", "_Month"])["Balance"]
        .last()
        .unstack(level="_Month", fill_value=None)
        .reindex(columns=all_months)  # includes months with no transactions
    )
    # Forward-fill so months without transactions carry the prior month's balance
    grp = grp.ffill(axis=1).fillna(0.0)

    # Sum by AllocBD
    alloc_bal = grp.groupby(level="_AllocBD").sum()

    monthly: dict[str, dict[str, float]] = {}
    for alloc_bd in alloc_bal.index:
        monthly[alloc_bd] = {m: float(alloc_bal.at[alloc_bd, m]) for m in all_months}

    return monthly, all_months


def _compute_rv(monthly: dict, periods: list[str], is_quarter_or_year: bool = False,
                quarter_last_month: dict | None = None) -> dict[str, dict[str, float]]:
    """
    Build {label: {period: value}} from the monthly balances and a list of target periods.

    If is_quarter_or_year=True, each period maps to a specific month via quarter_last_month.
    """
    def v(alloc_bd: str, period: str) -> float:
        if is_quarter_or_year and quarter_last_month:
            m = quarter_last_month.get(period)
            if m is None:
                return 0.0
            return monthly.get(alloc_bd, {}).get(m, 0.0)
        return monthly.get(alloc_bd, {}).get(period, 0.0)

    rv: dict[str, dict[str, float]] = {}

    # ── Data rows ─────────────────────────────────────────────────────────────
    for label, rtype, extra in BS_BD_ROWS:
        if rtype == "data":
            rv[label] = {p: v(extra, p) for p in periods}

    # ── Subtotals ─────────────────────────────────────────────────────────────
    for label, rtype, extra in BS_BD_ROWS:
        if rtype == "subtotal":
            rv[label] = {
                p: sum(rv.get(ch, {}).get(p, 0.0) for ch in extra if ch != "Retained earnings")
                for p in periods
            }

    # ── Totals (multi-pass for chain deps) ────────────────────────────────────
    total_rows = [(lbl, ext) for lbl, rt, ext in BS_BD_ROWS if rt == "total"]
    for _ in range(4):
        for label, extra in total_rows:
            if all(dep in rv for dep in extra):
                rv[label] = {
                    p: sum(rv.get(dep, {}).get(p, 0.0) for dep in extra)
                    for p in periods
                }

    # ── Retained earnings plug ────────────────────────────────────────────────
    equity_excl = ["Common stock", "Preferred Series B", "Preferred Series B-1",
                   "Preferred Series B-2", "Additional paid in capital"]
    rv["Retained earnings"] = {
        p: (
            rv.get("Total Assets", {}).get(p, 0.0)
            - rv.get("Total Current Liabilities", {}).get(p, 0.0)
            - rv.get("Non-current liabilities", {}).get(p, 0.0)
            - sum(rv.get(eq, {}).get(p, 0.0) for eq in equity_excl)
        )
        for p in periods
    }

    # Recompute Total stockholders' equity and Total L&SE with retained earnings
    se_children = ["Common stock", "Preferred Series B", "Preferred Series B-1",
                   "Preferred Series B-2", "Additional paid in capital", "Retained earnings"]
    rv["Total stockholders' equity"] = {
        p: sum(rv.get(ch, {}).get(p, 0.0) for ch in se_children)
        for p in periods
    }
    rv["Total liabilities and stockholders' equity"] = {
        p: (rv.get("Total Current Liabilities", {}).get(p, 0.0)
            + rv.get("Non-current liabilities", {}).get(p, 0.0)
            + rv.get("Total stockholders' equity", {}).get(p, 0.0))
        for p in periods
    }

    # ── Check = Total Assets − Total L&SE ─────────────────────────────────────
    rv["Check"] = {
        p: rv.get("Total Assets", {}).get(p, 0.0)
           - rv.get("Total liabilities and stockholders' equity", {}).get(p, 0.0)
        for p in periods
    }

    return rv


# ── Quarter/year helpers ──────────────────────────────────────────────────────

def _quarters_in_year(year: int, all_months: list[str]) -> list[str]:
    """Return Q1–Q4 of the given year, filtered to quarters that have data."""
    available_qs = {_month_to_quarter(m) for m in all_months}
    return [f"Q{n}-{year}" for n in range(1, 5) if f"Q{n}-{year}" in available_qs]


def _last_month_of_quarter(q: str, all_months: list[str]) -> str | None:
    """Return the strict quarter-end month: Q1→Mar, Q2→Jun, Q3→Sep, Q4→Dec."""
    try:
        q_num = int(q[1])
        year  = int(q[3:])
    except (ValueError, IndexError):
        return None
    end_month  = q_num * 3  # Q1→3, Q2→6, Q3→9, Q4→12
    target_ym  = f"{year}-{str(end_month).zfill(2)}"
    for m in all_months:
        if _month_to_ym(m) == target_ym:
            return m
    return None


def _last_month_of_year(year: int, all_months: list[str]) -> str | None:
    """Return the last month we have data for within the given year."""
    for m_num in range(12, 0, -1):
        target_ym = f"{year}-{str(m_num).zfill(2)}"
        for m in reversed(all_months):
            if _month_to_ym(m) == target_ym:
                return m
    return None


# ── Aggregate ─────────────────────────────────────────────────────────────────

def _aggregate(df: pd.DataFrame) -> dict:
    """
    Returns a preview-dict with quarterly and annual computed values.

    Shape:
      {
        available_quarters: ["Q1-2025", ...],
        available_years:    [2024, 2025],
        data: {
          "Q1-2025": { alloc_bd_label: value, ... },
          "2024":    { ... },
          ...
        },
        rows: [ {label, type, key}, ... ]
      }
    """
    monthly, all_months = _build_monthly_balances(df)
    if not all_months:
        return {
            "available_quarters": [], "available_years": [], "data": {}, "rows": _row_defs(),
        }

    years = sorted({int(_month_to_ym(m)[:4]) for m in all_months})
    all_quarters = _sort_quarters(list({_month_to_quarter(m) for m in all_months}))

    # Build quarter-last-month map
    q_last: dict[str, str] = {}
    for q in all_quarters:
        lm = _last_month_of_quarter(q, all_months)
        if lm:
            q_last[q] = lm

    # Build year-last-month map
    y_last: dict[str, str] = {}
    for yr in years:
        lm = _last_month_of_year(yr, all_months)
        if lm:
            y_last[str(yr)] = lm

    all_periods = list(q_last.keys()) + list(y_last.keys())
    combined_last = {**q_last, **y_last}

    rv = _compute_rv(monthly, all_periods, is_quarter_or_year=True, quarter_last_month=combined_last)

    def _round(v: float) -> float | None:
        return round(v, 2) if v is not None else None

    data: dict[str, dict] = {}
    for period in all_periods:
        data[period] = {k: _round(vd.get(period, 0.0)) for k, vd in rv.items()}

    return {
        "available_quarters": _sort_quarters(list(q_last.keys())),
        "available_years":    years,
        "data":               data,
        "rows":               _row_defs(),
    }


def _row_defs() -> list[dict]:
    result = []
    for label, rtype, extra in BS_BD_ROWS:
        key = None
        if rtype in ("data", "subtotal", "total", "check"):
            key = label
        elif rtype == "retained":
            key = "Retained earnings"
        result.append({"label": label, "type": rtype, "key": key})
    return result


# ── Quarter helper (mirrors comp_pl_bd logic) ─────────────────────────────────

def _four_quarters_ending_at(q: str) -> list[str]:
    """Return the 4 consecutive quarters ending at q (inclusive), e.g. Q4-2025 → [Q1,Q2,Q3,Q4]-2025."""
    try:
        q_num = int(q[1])
        year  = int(q[3:])
    except (ValueError, IndexError):
        return []
    result = []
    for i in range(3, -1, -1):
        qi = q_num - i
        yi = year
        while qi <= 0:
            qi += 4; yi -= 1
        result.append(f"Q{qi}-{yi}")
    return result


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_bs_bd_excel(
    data: dict,
    period_cols: list[str],
    as_of_quarter: str,
    company_name: str,
    year_cols: list[str] | None = None,
) -> bytes:
    """Render BS (BD) Excel for the given quarterly period_cols + optional year_cols."""
    period_cols = [p for p in period_cols if p in data]
    year_cols   = [y for y in (year_cols or []) if y in data]
    all_cols    = period_cols + year_cols
    n_q = len(period_cols)
    n_y = len(year_cols)
    n   = len(all_cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "BS (BD)"

    if not all_cols:
        ws["A1"] = "No Balance Sheet data found."
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    # ── Styles ────────────────────────────────────────────────────────────────
    def F(bold=False, size=9, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    DARK_NAVY  = PatternFill("solid", fgColor="0F172A")
    MED_BLUE   = PatternFill("solid", fgColor="EBF2FB")
    LIGHT_GRAY = PatternFill("solid", fgColor="F1F5F9")
    SUBSEC_BG  = PatternFill("solid", fgColor="F8FAFC")
    GREEN_FILL = PatternFill("solid", fgColor="F0FDF4")

    AL = Alignment(horizontal="left",   vertical="center")
    AR = Alignment(horizontal="right",  vertical="center")
    AC = Alignment(horizontal="center", vertical="center")

    thin  = Side(border_style="thin",   color="CBD5E1")
    NUM   = "#,##0.00"

    DARK_PURPLE = PatternFill("solid", fgColor="400f61")

    # ── Title block ───────────────────────────────────────────────────────────
    ws.append([company_name] + [None] * n)
    ws["A1"].font = F(bold=True, size=14, color="0F172A"); ws.row_dimensions[1].height = 22

    ws.append(["Balance Sheets (BD)"] + [None] * n)
    ws["A2"].font = F(bold=True, size=11, color="1E3A5F")

    ws.append([f"As of {as_of_quarter}"] + [None] * n)
    ws["A3"].font = F(size=9, color="64748B", italic=True)

    ws.append([None] * (n + 1))  # blank row 4
    ws.append([None] * (n + 1))  # blank row 5 (freeze-pane alignment)

    # ── Period header row (row 6) ─────────────────────────────────────────────
    ws.append(["Particulars"] + all_cols)
    hr = ws.max_row
    ws.row_dimensions[hr].height = 16
    for ci in range(1, n + 2):
        c = ws.cell(hr, ci)
        c.fill = DARK_PURPLE
        c.font = F(bold=True, size=9, color="FFFFFF")
        c.alignment = AL if ci == 1 else AC

    # ── Data rows ─────────────────────────────────────────────────────────────
    for label, rtype, extra in BS_BD_ROWS:

        if rtype == "blank":
            ws.append([None] * (n + 1))
            continue

        key = None
        if rtype in ("data", "subtotal", "total", "check"):
            key = label
        elif rtype == "retained":
            key = "Retained earnings"

        vals = (
            [data.get(p, {}).get(key, 0.0) for p in all_cols]
            if key else [None] * n
        )

        display = label
        if rtype in ("data", "retained") and label:
            display = "  " + label

        ws.append([display] + vals)
        r = ws.max_row

        ca = ws.cell(r, 1)
        ca.alignment = AL

        def fill_row(fill, fill_year=None):
            for ci in range(1, n + 2):
                is_yr = fill_year and ci > n_q + 1
                ws.cell(r, ci).fill = fill_year if is_yr else fill

        for ci, v in enumerate(vals, 2):
            cell = ws.cell(r, ci)
            if v is not None:
                cell.number_format = NUM
                cell.alignment = AR

        if rtype == "section":
            fill_row(DARK_NAVY, DARK_NAVY)
            for ci in range(1, n + 2):
                ws.cell(r, ci).font = F(bold=True, size=9, color="FFFFFF")
            ws.row_dimensions[r].height = 15

        elif rtype == "subsection":
            fill_row(SUBSEC_BG, SUBSEC_BG)
            ca.font = F(bold=True, size=9, color="1E293B")

        elif rtype == "data":
            ca.font = F(size=9, color="334155")
            for ci in range(2, n + 2):
                c = ws.cell(r, ci)
                if c.value is not None:
                    neg = isinstance(c.value, (int, float)) and c.value < 0
                    c.font = F(size=9, color="DC2626" if neg else "334155")

        elif rtype == "retained":
            ca.font = F(size=9, color="334155")
            for ci in range(2, n + 2):
                c = ws.cell(r, ci)
                if c.value is not None:
                    neg = isinstance(c.value, (int, float)) and c.value < 0
                    c.font = F(size=9, color="DC2626" if neg else "334155")

        elif rtype == "subtotal":
            fill_row(MED_BLUE, MED_BLUE)
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.font   = F(bold=True, size=9, color="1E40AF")
                c.border = Border(top=thin)

        elif rtype == "total":
            fill_row(LIGHT_GRAY, LIGHT_GRAY)
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.font   = F(bold=True, size=9)
                c.border = Border(top=Side(border_style="medium", color="94A3B8"))
            ws.row_dimensions[r].height = 15

        elif rtype == "check":
            fill_row(GREEN_FILL, GREEN_FILL)
            for ci in range(1, n + 2):
                c = ws.cell(r, ci)
                c.font = F(bold=True, size=9, color="166534")

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 46
    for i in range(n):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14
    ws.freeze_panes = "B7"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Public API ────────────────────────────────────────────────────────────────

def run_bs_bd(df: pd.DataFrame, company_name: str) -> bytes:
    """Generate BS (BD) Excel from a raw DataFrame (defaults to last available quarter)."""
    preview  = _aggregate(df)
    data     = preview.get("data", {})
    quarters = preview.get("available_quarters", [])
    as_of    = quarters[-1] if quarters else ""
    period_cols = _four_quarters_ending_at(as_of) if as_of else []
    sel_year    = int(as_of[3:]) if as_of else 0
    year_cols   = [str(y) for y in [sel_year - 1, sel_year] if str(y) in data]
    return _build_bs_bd_excel(data, period_cols, as_of, company_name, year_cols)


def get_bs_bd_preview(df: pd.DataFrame, company_name: str = "") -> dict:
    """Return JSON-serialisable preview for the frontend."""
    result = _aggregate(df)
    result["company_name"] = company_name
    return result


def run_bs_bd_from_preview(
    preview: dict,
    selected_quarter: str | None,
    company_name: str,
) -> bytes:
    """Generate BS (BD) Excel from cached preview dict for the given quarter."""
    quarters = preview.get("available_quarters", [])
    data     = preview.get("data", {})
    as_of    = selected_quarter if selected_quarter else (quarters[-1] if quarters else "")
    period_cols = _four_quarters_ending_at(as_of) if as_of else []
    sel_year    = int(as_of[3:]) if as_of else 0
    year_cols   = [str(y) for y in [sel_year - 1, sel_year] if str(y) in data]
    return _build_bs_bd_excel(data, period_cols, as_of, company_name, year_cols)
