"""portco/parser.py — Parse MBR Excel files for PortCo Reporting dashboard."""
from __future__ import annotations
from io import BytesIO
from typing import Any

import openpyxl


def _safe_float(v: Any) -> float | None:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        return None if (f != f) else f  # NaN guard
    return None


def _fmt_date(dt: Any) -> str | None:
    if hasattr(dt, "year") and hasattr(dt, "month"):
        return f"{dt.year}-{str(dt.month).zfill(2)}"
    return None


def _find_label_row(rows: list, label: str, col: int = 1) -> int | None:
    for i, row in enumerate(rows):
        if len(row) > col and row[col] == label:
            return i
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Finance sheet
# ─────────────────────────────────────────────────────────────────────────────

def _parse_finance(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))

    # Company name / report month from first 4 rows
    company_name, report_month = "", ""
    for row in rows[:5]:
        for cell in row:
            if not isinstance(cell, str) or not cell.strip():
                continue
            s = cell.strip()
            if s in ("Finance", "Monthly"):
                continue
            if not company_name:
                company_name = s
            elif not report_month and "Reporting" in s:
                report_month = s

    # Date header row — find the row with the most datetime objects
    date_col_start = 4   # col E (0-indexed)
    dates: list = []
    for row in rows:
        d = [v for v in row[date_col_start:] if hasattr(v, "year")]
        if len(d) > len(dates):
            dates = d

    if not dates:
        return {"months": [], "company_name": company_name, "report_month": report_month}

    months = [_fmt_date(d) for d in dates]
    n = len(months)

    def series(label: str) -> list[float | None]:
        idx = _find_label_row(rows, label)
        if idx is None:
            return [None] * n
        row = rows[idx]
        vals = list(row[date_col_start: date_col_start + n])
        return [_safe_float(v) for v in vals]

    # Gross margin = Revenue + COGS (COGS are negative in the sheet)
    revenue  = series("Revenue")
    cogs_svc = series("(-) Service/Delivery (CoS/CoGS)")
    cogs_prd = series("(-) Product")
    cogs_tch = series("(-) Technology")
    gross_profit = [
        (r or 0) + (s or 0) + (p or 0) + (t or 0)
        for r, s, p, t in zip(revenue, cogs_svc, cogs_prd, cogs_tch)
    ]
    gross_margin_pct = [
        (gp / r * 100) if (r and r != 0) else None
        for gp, r in zip(gross_profit, revenue)
    ]
    ebitda = series("Adj. EBITDA")
    ebitda_margin_pct = [
        (e / r * 100) if (r and r != 0) else None
        for e, r in zip(ebitda, revenue)
    ]
    cash_raw = series("(-) Cash")
    cash = [abs(v) if v is not None else None for v in cash_raw]

    return {
        "months": months,
        "company_name": company_name,
        "report_month": report_month,
        "income_statement": {
            "ARR":            series("ARR"),
            "Revenue":        revenue,
            "COGS_Service":   cogs_svc,
            "COGS_Product":   cogs_prd,
            "COGS_Technology":cogs_tch,
            "Exp_Sales":      series("(-) Sales"),
            "Exp_Marketing":  series("(-) Marketing"),
            "Exp_CS":         series("(-) Customer Success"),
            "Exp_GA":         series("(-) G&A"),
            "GrossProfit":    gross_profit,
            "GrossMarginPct": gross_margin_pct,
            "EBITDA":         ebitda,
            "EBITDAMarginPct":ebitda_margin_pct,
        },
        "balance_sheet": {
            "Cash":    cash,
            "NetDebt": series("Net Debt"),
            "Debt":    series("Debt & Debt-like Items"),
        },
        "headcount": {
            "Employees":   series("Employees"),
            "VeArc":       series("VeArc"),
            "ThirdParty":  series("3rd Party Providers"),
        },
        "rule_200": {
            "NewLogoGrowth":    series("New Logo ARR Growth (annualized)"),
            "NRR":              series("NRR (annualized)"),
            "GrossMargin":      series("Gross Margin"),
            "EBITDAMargin":     series("EBITDA Margin"),
            "Score":            series("Score"),
            "NewLogoGrowthYTD": series("New Logo ARR Growth (YTD)"),
            "NRRYTD":           series("NRR (YTD)"),
            "GrossMarginYTD":   series("Gross Margin (YTD)"),
            "EBITDAMarginYTD":  series("EBITDA Margin (YTD)"),
            "ScoreYTD":         series("Score (YTD)"),
        },
        "zero_based": {
            "AvgPersonnelCostPerSquad": series("Avg. Personnel Cost per Squad"),
            "PTOverhead":               series("P&T Overhead"),
            "MktgExpPerARR":            series("Mktg exp. per $1 ARR"),
            "SalesExpPerARR":           series("Sales exp. per $1 ARR"),
            "CSExpPerARR":              series("C.S. exp. per $1 ARR"),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Actuals / Budget sheet (same structure)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_actuals(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains 'Department' in col B)
    header_idx = None
    for i, row in enumerate(rows):
        if len(row) > 1 and row[1] == "Department":
            header_idx = i
            break

    if header_idx is None:
        return {"months": [], "data": {}}

    header = rows[header_idx]
    date_col_start = 7   # col H (Dept, Cat, ID, Line Item, Units, Assignee = 6 cols after col A)
    dates = [v for v in header[date_col_start:] if hasattr(v, "year")]
    months = [_fmt_date(d) for d in dates]
    n = len(months)

    data: dict[str, dict[str, list]] = {}
    for row in rows[header_idx + 1:]:
        dept = row[1] if len(row) > 1 else None
        line_item = row[4] if len(row) > 4 else None
        if not dept or not line_item:
            continue
        vals = list(row[date_col_start: date_col_start + n])
        if dept not in data:
            data[dept] = {}
        data[dept][line_item] = [_safe_float(v) for v in vals]

    return {"months": months, "data": data}


# ─────────────────────────────────────────────────────────────────────────────
# Executive Summary sheet
# ─────────────────────────────────────────────────────────────────────────────

def _parse_exec_summary(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    results = []
    skip = {"", "Concertiv", "Executive Summary", "What Went Well",
            "What Did Not Go Well", "Focus Areas for next month"}
    for row in rows:
        dept = row[1] if len(row) > 1 else None
        if not dept or not isinstance(dept, str):
            continue
        dept = dept.strip()
        if dept in skip or "Reporting" in dept:
            continue
        went_well = str(row[2] or "").strip() if len(row) > 2 else ""
        not_well  = str(row[3] or "").strip() if len(row) > 3 else ""
        focus     = str(row[4] or "").strip() if len(row) > 4 else ""
        if went_well or not_well or focus:
            results.append({
                "section":    dept,
                "went_well":  went_well,
                "not_well":   not_well,
                "focus":      focus,
            })
    return results


# ─────────────────────────────────────────────────────────────────────────────
# MetricMap parser — returns {metric_id: {YYYY-MM: value}}
# metric_id = "{Department} {Line Item}"  (matches frontend METRIC_DEFS ids)
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_month_str(v: Any) -> str | None:
    """Try to extract YYYY-MM from a datetime, date, or string value."""
    if hasattr(v, "year") and hasattr(v, "month"):
        return f"{v.year}-{str(v.month).zfill(2)}"
    if isinstance(v, str):
        import re
        # e.g. "Jan-2024", "January 2024", "2024-01"
        m = re.match(r"(\d{4})-(\d{2})", v.strip())
        if m:
            return v.strip()[:7]
        m = re.search(r"(\d{4})", v)
        if m:
            month_names = ["jan","feb","mar","apr","may","jun",
                           "jul","aug","sep","oct","nov","dec"]
            lo = v.lower()
            for idx, mn in enumerate(month_names):
                if mn in lo:
                    return f"{m.group(1)}-{str(idx+1).zfill(2)}"
    return None


def _find_sheet(wb, name: str):
    """Case-insensitive sheet lookup."""
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.lower():
            return wb[sn]
    return None


def _sheet_to_metric_map(ws) -> dict[str, dict[str, float]]:
    """Convert an Actuals or Budget sheet to MetricMap format."""
    rows = list(ws.iter_rows(values_only=True))

    # Find header row: contains "Department" in col B (index 1)
    header_idx = None
    for i, row in enumerate(rows):
        if len(row) > 1 and isinstance(row[1], str) and row[1].strip() == "Department":
            header_idx = i
            break
    if header_idx is None:
        return {}

    header = rows[header_idx]

    # Detect where date columns begin: first cell that parses to a YYYY-MM string
    date_col_start = None
    for j, v in enumerate(header):
        if _fmt_month_str(v) is not None:
            date_col_start = j
            break
    # Fallback to column 7 (col H) — the hardcoded position from _parse_actuals
    if date_col_start is None:
        date_col_start = 7

    months = [_fmt_month_str(v) for v in header[date_col_start:]]
    months = [m for m in months if m]  # drop Nones
    n = len(months)
    if n == 0:
        return {}

    metric_map: dict[str, dict[str, float]] = {}
    for row in rows[header_idx + 1:]:
        dept      = str(row[1]).strip() if len(row) > 1 and row[1] else None
        line_item = str(row[4]).strip() if len(row) > 4 and row[4] else None
        if not dept or not line_item:
            continue

        metric_id = f"{dept} {line_item}"
        # Collect values aligned to the detected month positions
        month_vals = list(row[date_col_start: date_col_start + n])

        month_map: dict[str, float] = {}
        for month, val in zip(months, month_vals):
            f = _safe_float(val)
            if f is not None:
                month_map[month] = f

        if month_map:
            metric_map[metric_id] = month_map

    return metric_map


def _flat_sheet_to_metric_map(ws) -> dict[str, dict[str, float]]:
    """Parse a flat export sheet (Department, Category, ID, Line Item, Units, [months...]).
    The ID column (index 2) already contains the metric_id.
    Date columns start at index 5.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = rows[0]
    # Dates start at col 5; detect actual positions
    date_col_start = next(
        (j for j in range(5, len(header)) if _fmt_month_str(header[j]) is not None),
        5,
    )
    months = [_fmt_month_str(header[j]) for j in range(date_col_start, len(header))]
    months = [m for m in months if m]
    n = len(months)
    if n == 0:
        return {}

    metric_map: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        dept      = str(row[0]).strip() if len(row) > 0 and row[0] else None
        line_item = str(row[3]).strip() if len(row) > 3 and row[3] else None
        if not dept or not line_item:
            continue

        # Prefer pre-computed ID col (plain string); fall back to Dept + Line Item
        # when the ID cell is a formula whose cache was cleared by openpyxl.
        id_cell = row[2] if len(row) > 2 else None
        if id_cell and isinstance(id_cell, str) and id_cell.strip() and not id_cell.startswith("="):
            metric_id = id_cell.strip()
        else:
            metric_id = f"{dept} {line_item}"

        month_vals = list(row[date_col_start: date_col_start + n])
        month_map: dict[str, float] = {}
        for month, val in zip(months, month_vals):
            f = _safe_float(val)
            if f is not None:
                month_map[month] = f

        if month_map:
            metric_map[metric_id] = month_map

    return metric_map


def _is_flat_format(wb) -> bool:
    """Return True if the workbook uses the flat export format (Sheet1 with ID column)."""
    ws = _find_sheet(wb, "Sheet1")
    if ws is None:
        return False
    rows = list(ws.iter_rows(values_only=True, max_row=2))
    if not rows:
        return False
    header = rows[0]
    # Flat format has 'ID' at col 2
    return len(header) > 2 and isinstance(header[2], str) and header[2].strip() == "ID"


def parse_mbr_to_metric_map(file_bytes: bytes) -> dict:
    """Parse an MBR Excel file and return {actuals: MetricMap, budget: MetricMap}.

    Supports two formats:
    - Flat export (Actuals.xlsx / Budget.xlsx): single Sheet1 with columns
      Department, Category, ID, Line Item, Units, [months...]. The file
      contains data for one mode (actuals or budget); the caller decides which.
    - MBR dashboard (multi-sheet): has separate 'Actuals' and 'Budget' sheets.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    if _is_flat_format(wb):
        ws = _find_sheet(wb, "Sheet1")
        data = _flat_sheet_to_metric_map(ws)
        # Return data under both keys; the caller (UploadZone) uses only the
        # relevant one based on which button was clicked.
        return {"actuals": data, "budget": data}

    # MBR multi-sheet format
    actuals_ws = _find_sheet(wb, "Actuals")
    budget_ws  = _find_sheet(wb, "Budget")
    actuals = _sheet_to_metric_map(actuals_ws) if actuals_ws is not None else {}
    budget  = _sheet_to_metric_map(budget_ws)  if budget_ws  is not None else {}
    return {"actuals": actuals, "budget": budget}


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def parse_mbr_file(file_bytes: bytes) -> dict:
    """Parse an MBR Excel workbook and return structured dashboard data."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    finance = _parse_finance(wb["Finance"]) if "Finance" in wb.sheetnames else {}
    company_name = finance.pop("company_name", "")
    report_month = finance.pop("report_month", "")

    actuals = _parse_actuals(wb["Actuals"]) if "Actuals" in wb.sheetnames else {"months": [], "data": {}}
    budget  = _parse_actuals(wb["Budget"])  if "Budget"  in wb.sheetnames else {"months": [], "data": {}}
    exec_summary = (
        _parse_exec_summary(wb["Executive_Summary"])
        if "Executive_Summary" in wb.sheetnames else []
    )

    return {
        "company_name":  company_name,
        "report_month":  report_month,
        "finance":       finance,
        "actuals":       actuals,
        "budget":        budget,
        "exec_summary":  exec_summary,
    }
