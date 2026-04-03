# =============================================================================
# processing/consolidator.py  —  Append each payroll run to master files
# =============================================================================
"""
Two consolidated files are maintained in /consolidated:

  Consolidated_Payroll.xlsx  — one JE per cycle (updated on Download)
  Consolidated_Inputs.xlsx   — raw payroll rows per cycle (updated on Generate)

Layout of each file
-------------------
  Row 1        : Column headers (dark green)
  (blank row)
  Cycle banner : "Payroll Cycle: Salary for 01/30/2026"  (light green, merged)
  Data rows    : alternating white / very-light-green
  (blank row)
  Cycle banner : next cycle …

Re-running the same cycle replaces its block so there are no duplicates.
"""

from __future__ import annotations

from pathlib import Path
from typing import List
import threading

import pandas as pd
from openpyxl            import load_workbook, Workbook
from openpyxl.styles     import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils      import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from config import JE_COLUMNS

# ── Paths ────────────────────────────────────────────────────────────────────
_BASE_DIR             = Path(__file__).resolve().parent.parent
CONSOLIDATED_DIR      = _BASE_DIR / "consolidated"
CONSOLIDATED_PATH     = CONSOLIDATED_DIR / "Consolidated_Payroll.xlsx"
CONSOLIDATED_INPUTS_PATH = CONSOLIDATED_DIR / "Consolidated_Inputs.xlsx"

# ── Write lock — prevents two threads writing to the same file simultaneously ─
_JE_LOCK    = threading.Lock()
_INPUT_LOCK = threading.Lock()

# ── Palette ──────────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1E7E34")   # deep green  – column headers
_CYCLE_FILL = PatternFill("solid", fgColor="C6EFCE")   # pale green  – cycle banner
_ROW_ALT    = PatternFill("solid", fgColor="F2F9F3")   # very pale   – alternating rows
_ROW_WHITE  = PatternFill("solid", fgColor="FFFFFF")   # white rows

# ── Fonts ────────────────────────────────────────────────────────────────────
_HDR_FONT   = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
_CYCLE_FONT = Font(name="Calibri", bold=True,  color="1E7E34", size=10)
_DATA_FONT  = Font(name="Calibri", bold=False, color="000000", size=10)

# ── Alignments ───────────────────────────────────────────────────────────────
_ALIGN_HDR   = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
_ALIGN_LEFT  = Alignment(horizontal="left",   vertical="center",
                          wrap_text=False)   # NO wrap on data – keeps rows single-height
_ALIGN_RIGHT = Alignment(horizontal="right",  vertical="center",
                          wrap_text=False)
_ALIGN_CYCLE = Alignment(horizontal="left",   vertical="center",
                          wrap_text=False, indent=1)

# ── Borders ──────────────────────────────────────────────────────────────────
_S   = Side(style="thin",   color="BFBFBF")
_SB  = Side(style="medium", color="1E7E34")   # bottom border on header

_BORDER_HDR  = Border(left=_S, right=_S, top=_SB, bottom=_SB)
_BORDER_DATA = Border(left=_S, right=_S, top=_S,  bottom=_S)
_BORDER_CYCLE= Border(left=_SB, right=_SB, top=_SB, bottom=_SB)

# ── Column widths for the JE consolidated file ───────────────────────────────
# Keyed by JE_COLUMNS index (0-based) → width in characters
_JE_COL_WIDTHS: List[float] = [
    7,    # Post?
    28,   # Journal Number
    13,   # Entry Date
    45,   # Journal Description
    70,   # Account  ← widened to fit longest GL string (68 chars)
    14,   # Account ID
    14,   # Customer
    38,   # Vendor
    28,   # Employee
    14,   # Location
    36,   # Class
    12,   # Tax Rate
    20,   # Tax Application ON
    10,   # Currency
    17,   # Debit (exc. Tax)
    17,   # Credit (exc. Tax)
    12,   # Adjustment
    14,   # QBO Edit ID
]

# ── Row heights ───────────────────────────────────────────────────────────────
_H_HEADER = 36    # column-header row
_H_CYCLE  = 22    # cycle-banner row
_H_DATA   = 20    # data rows — single-line; no wrapping
_H_BLANK  = 10    # separator blank row


# =============================================================================
# PUBLIC API
# =============================================================================

def append_to_consolidated(je_df: pd.DataFrame, journal_number: str) -> Path:
    """Append / replace one JE cycle in Consolidated_Payroll.xlsx."""
    with _JE_LOCK:
        CONSOLIDATED_DIR.mkdir(parents=True, exist_ok=True)

        if CONSOLIDATED_PATH.exists():
            wb = load_workbook(CONSOLIDATED_PATH)
            ws = wb.active
            _remove_duplicate_headers(ws)   # fix any stray header rows from prior runs
            _remove_cycle(ws, journal_number)
        else:
            wb, ws = _new_workbook("Consolidated Payroll JE")
            _write_je_headers(ws)
            _set_je_col_widths(ws)
            ws.freeze_panes = "A2"

        _blank_separator(ws, len(JE_COLUMNS))
        _cycle_banner(ws, journal_number, len(JE_COLUMNS))
        _write_je_rows(ws, je_df)

        wb.save(CONSOLIDATED_PATH)
        return CONSOLIDATED_PATH


def append_input_to_consolidated(raw_file_bytes: bytes, journal_number: str) -> Path:
    """
    Append one payroll cycle to Consolidated_Inputs.xlsx.

    Performs a true copy of the source Excel file into the consolidated sheet:
      - Cell values, number formats, fonts, fills, borders, alignments
      - Row heights and column widths (applied once from the first cycle)
      - Merged cells (offsets adjusted for position in the consolidated sheet)
      - Copyright footer row (©) is excluded

    Re-running the same cycle removes the old block first (idempotent).
    """
    from io import BytesIO
    from copy import copy as _copy
    import openpyxl as _oxl

    CONSOLIDATED_DIR.mkdir(parents=True, exist_ok=True)

    # ── 1. Open source with full formatting (data_only=True for evaluated values) ──
    src_wb = _oxl.load_workbook(BytesIO(raw_file_bytes), data_only=True)
    src_ws = src_wb.active

    # Find the last row to copy (stop before copyright footer)
    src_last_row = 0
    for row in src_ws.iter_rows(min_col=1, max_col=1):
        val = str(row[0].value or "").strip()
        if val.startswith("©") or val.startswith("\u00a9"):
            break
        src_last_row = row[0].row

    if src_last_row == 0:
        src_wb.close()
        return CONSOLIDATED_INPUTS_PATH

    # Thin border applied to every data cell so the file looks like a spreadsheet
    _THIN  = Side(style="thin", color="BFBFBF")
    _CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    # ── 2. Load or create destination workbook ────────────────────────────────
    if CONSOLIDATED_INPUTS_PATH.exists():
        dst_wb = load_workbook(CONSOLIDATED_INPUTS_PATH)
        dst_ws = dst_wb.active
        _remove_input_cycle(dst_ws, journal_number)
        # One blank separator row between cycles (only when there is prior content)
        if dst_ws.max_row > 0 and dst_ws.cell(dst_ws.max_row, 1).value is not None:
            dst_ws.append([None] * src_ws.max_column)
        is_first = False
        # row_offset set AFTER any separator so the new cycle starts at the right row
        row_offset = dst_ws.max_row
    else:
        dst_wb = Workbook()
        dst_ws = dst_wb.active
        dst_ws.title = "Consolidated Inputs"
        dst_ws.sheet_view.showGridLines = True   # keep Excel grid visible
        is_first = True
        row_offset = 0   # fresh sheet — first source row goes to Excel row 1

    # ── 3. Copy column widths once (from the first cycle) ────────────────────
    if is_first:
        for col_letter, col_dim in src_ws.column_dimensions.items():
            if col_dim.width:
                dst_ws.column_dimensions[col_letter].width = col_dim.width
        # Enforce minimum width on every column so no cell ever shows ########
        # First pass: columns already tracked by openpyxl
        for col_letter, col_dim in dst_ws.column_dimensions.items():
            if (col_dim.width or 0) < 10:
                dst_ws.column_dimensions[col_letter].width = 10
        # Second pass: columns used in the data but not explicitly dimensioned
        for col_idx in range(1, src_ws.max_column + 1):
            ltr = get_column_letter(col_idx)
            if (dst_ws.column_dimensions[ltr].width or 0) < 10:
                dst_ws.column_dimensions[ltr].width = 10

    # ── 4. Copy every row: values + full cell formatting + row height ─────────
    import datetime as _dt
    for src_rn in range(1, src_last_row + 1):
        dst_rn = row_offset + src_rn

        # Row height
        src_rd = src_ws.row_dimensions.get(src_rn)
        if src_rd and src_rd.height:
            dst_ws.row_dimensions[dst_rn].height = src_rd.height

        for src_cell in src_ws[src_rn]:
            dst_cell = dst_ws.cell(row=dst_rn, column=src_cell.column)
            dst_cell.value = src_cell.value

            # Copy source formatting
            if src_cell.has_style:
                dst_cell.font          = _copy(src_cell.font)
                dst_cell.fill          = _copy(src_cell.fill)
                dst_cell.alignment     = _copy(src_cell.alignment)
                # For date/datetime cells, standardise the number format to MM/DD/YYYY
                # so dates always display with slashes regardless of source format
                if isinstance(src_cell.value, (_dt.date, _dt.datetime)):
                    dst_cell.number_format = "MM/DD/YYYY"
                else:
                    dst_cell.number_format = src_cell.number_format
                # Keep source border if it has one; otherwise apply thin grid border
                if src_cell.border and src_cell.border != Border():
                    dst_cell.border = _copy(src_cell.border)
                else:
                    dst_cell.border = _CELL_BORDER
            else:
                # Cell has no style at all — add thin border so it's part of the grid
                # Still normalise date format even when there is no other style
                if isinstance(src_cell.value, (_dt.date, _dt.datetime)):
                    dst_cell.number_format = "MM/DD/YYYY"
                dst_cell.border = _CELL_BORDER

    # ── 5. Copy merged cells (shift row indices by row_offset) ───────────────
    for merge in src_ws.merged_cells.ranges:
        if merge.max_row <= src_last_row:
            try:
                dst_ws.merge_cells(
                    start_row    = row_offset + merge.min_row,
                    start_column = merge.min_col,
                    end_row      = row_offset + merge.max_row,
                    end_column   = merge.max_col,
                )
            except Exception:
                pass

    src_wb.close()
    dst_wb.save(CONSOLIDATED_INPUTS_PATH)
    return CONSOLIDATED_INPUTS_PATH


# =============================================================================
# SHARED HELPERS
# =============================================================================

def _new_workbook(sheet_title: str):
    wb = Workbook()
    ws = wb.active
    ws.title        = sheet_title
    ws.sheet_view.showGridLines = False   # cleaner without grid
    return wb, ws


def _remove_duplicate_headers(ws) -> None:
    """
    Remove any header rows that appear after row 1.
    Identifies header rows by matching the first JE column header value.
    Also removes blank rows immediately preceding a duplicate header.
    """
    expected = JE_COLUMNS[0].replace(" (exc. Tax)", "")   # "Post?"
    to_del: list[int] = []
    for row in ws.iter_rows(min_row=2):
        rn  = row[0].row
        val = str(row[0].value or "").strip()
        if val == expected:
            # Remove the blank separator row right before this duplicate header (if any)
            if rn > 2 and ws.cell(rn - 1, 1).value in (None, ""):
                to_del.append(rn - 1)
            to_del.append(rn)
    for r in reversed(sorted(set(to_del))):
        ws.delete_rows(r)


def _remove_cycle(ws, journal_number: str) -> None:
    """Delete all rows belonging to an existing cycle (idempotent)."""
    target   = f"Payroll Cycle: {journal_number}"
    to_del: list[int] = []
    inside   = False

    for row in ws.iter_rows():
        rn        = row[0].row
        cell_val  = str(row[0].value or "").strip()

        if cell_val == target:
            inside = True

        if inside:
            to_del.append(rn)
            # end of block = blank row after at least one data row
            if len(to_del) > 2 and cell_val == "":
                inside = False

    for r in reversed(to_del):
        ws.delete_rows(r)


def _blank_separator(ws, n_cols: int) -> None:
    """Insert a thin blank row between cycles (skip if sheet is fresh)."""
    if ws.max_row < 1:
        return
    last = ws.cell(row=ws.max_row, column=1).value
    if last is not None and str(last).strip() != "":
        ws.append([""] * n_cols)
        ws.row_dimensions[ws.max_row].height = _H_BLANK


def _cycle_banner(ws, journal_number: str, n_cols: int) -> None:
    """Write a merged, styled cycle-header banner row."""
    label = f"Payroll Cycle: {journal_number}"
    ws.append([label] + [""] * (n_cols - 1))
    rn       = ws.max_row
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{rn}:{last_col}{rn}")
    ws.row_dimensions[rn].height = _H_CYCLE

    cell            = ws.cell(row=rn, column=1)
    cell.value      = label
    cell.fill       = _CYCLE_FILL
    cell.font       = _CYCLE_FONT
    cell.alignment  = _ALIGN_CYCLE
    cell.border     = _BORDER_CYCLE


# =============================================================================
# JE-SPECIFIC HELPERS
# =============================================================================

def _write_je_headers(ws) -> None:
    # Use short display names (remove "(exc. Tax)" from header text for brevity)
    headers = []
    for c in JE_COLUMNS:
        h = c.replace(" (exc. Tax)", "")
        headers.append(h)

    ws.append(headers)
    rn = ws.max_row
    ws.row_dimensions[rn].height = _H_HEADER

    for ci, _ in enumerate(headers, start=1):
        cell            = ws.cell(row=rn, column=ci)
        cell.fill       = _HDR_FILL
        cell.font       = _HDR_FONT
        cell.alignment  = _ALIGN_HDR
        cell.border     = _BORDER_HDR


def _set_je_col_widths(ws) -> None:
    for i, w in enumerate(_JE_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_je_rows(ws, je_df: pd.DataFrame) -> None:
    numeric_cols = {"Debit (exc. Tax)", "Credit (exc. Tax)"}

    for i, (_, row_data) in enumerate(je_df.iterrows()):
        vals = []
        for col in JE_COLUMNS:
            v = row_data.get(col, None)
            if not isinstance(v, str) and pd.isna(v):
                v = None
            vals.append(v)

        ws.append(vals)
        rn   = ws.max_row
        fill = _ROW_ALT if i % 2 else _ROW_WHITE
        ws.row_dimensions[rn].height = _H_DATA

        for ci, col_name in enumerate(JE_COLUMNS, start=1):
            cell           = ws.cell(row=rn, column=ci)
            cell.font      = _DATA_FONT
            cell.border    = _BORDER_DATA
            cell.fill      = fill

            if col_name in numeric_cols:
                cell.alignment     = Alignment(horizontal="right",  vertical="center", wrap_text=False)
                cell.number_format = '#,##0.00'
            else:
                cell.alignment     = Alignment(horizontal="left",   vertical="center", wrap_text=False)


# =============================================================================
# INPUT-SPECIFIC HELPERS
# =============================================================================


def _remove_input_cycle(ws, journal_number: str) -> None:
    """
    Delete the rows for an existing input cycle identified by its date.

    Each cycle block starts with "Invoice Supporting Details" in column A.
    We find the block whose Payroll Cycle row contains the date extracted
    from journal_number (e.g. "Salary for 01/30/2026" → "01/30/2026"),
    then delete from that block's start to just before the next block.
    """
    import re as _re
    date_m = _re.search(r'\d{1,2}/\d{1,2}/\d{4}', journal_number)
    if not date_m:
        return
    cycle_date = date_m.group(0)

    # Collect all row numbers where a new cycle block starts
    block_starts: list[int] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        rn  = row[0].row
        val = str(row[0].value or "").strip()
        if val == "Invoice Supporting Details":
            block_starts.append(rn)

    # Find which block contains our cycle_date (check up to 5 rows after start)
    target_start: int | None = None
    for s_rn in block_starts:
        for offset in range(6):
            cell_val = str(ws.cell(s_rn + offset, 1).value or "")
            if cycle_date in cell_val:
                target_start = s_rn
                break
        if target_start:
            break

    if target_start is None:
        return

    # Block ends just before the next block starts (or at end of sheet)
    target_end = ws.max_row
    for s_rn in block_starts:
        if s_rn > target_start:
            target_end = s_rn - 1   # -1 to include blank separator row before next block
            break

    # Include the blank separator row that precedes this block (if any)
    delete_from = target_start
    if delete_from > 1 and ws.cell(delete_from - 1, 1).value is None:
        delete_from -= 1

    for r in range(target_end, delete_from - 1, -1):
        ws.delete_rows(r)


